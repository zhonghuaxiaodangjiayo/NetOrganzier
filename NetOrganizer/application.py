import statistics
import openpyxl
from statsmodels.stats import multitest
import seaborn as sns
import os

import importdatas as datas
from sklearn.metrics.pairwise import cosine_similarity
from multiprocessing import Process, Queue, Manager
import random
import pickle
import heapq
from collections import OrderedDict
import numpy as np
import pandas as pd

from sklearn.metrics import precision_recall_curve

from sklearn import preprocessing

import matplotlib
import matplotlib.pyplot as plt

from scipy import stats
from scipy.stats import gaussian_kde
import scipy.stats
from scipy.stats import levene

from scipy.stats import hypergeom
from statsmodels.stats.multitest import multipletests
from statsmodels.stats.multitest import fdrcorrection
import csv


def calSimAvg():
    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}
    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    vecs = importvec('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/4Bert/3_27_regoaPath_4layerTrain_humanGRN_vec.txt')
    # vecs = importvec('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/vec/10_15_7path_addGB140w_epoch100_4layerTrain_humanGRN_2_vec.txt')

    # 初始化结果列表
    gene_results = []
    re_results = []

    # 遍历每一个组织
    for organ in geneORGANizer['bodyparts'].keys(): 

        # 计算该组织与每个基因的余弦相似度，并加入结果列表
        gene_similarities = []
        for gene in total_geneset:
            if gene in vecs:
                gene_similarities.append(cosine_similarity([vecs[gene], vecs[organ]])[0][1])
        gene_results.append(np.mean(gene_similarities))

        # 计算该组织与每个调控元件的余弦相似度，并加入结果列表
        re_similarities = []
        for re in grn['res'].keys():
            if re in vecs:
                re_similarities.append(cosine_similarity([vecs[re], vecs[organ]])[0][1])
        re_results.append(np.mean(re_similarities))

    # 可视化比较基因和调控元件在每个组织中的余弦相似度
    fig, ax = plt.subplots()
    x = np.arange(len(geneORGANizer['bodyparts']))
    width = 0.35

    rects1 = ax.bar(x - width/2, gene_results, width, label='Gene')
    rects2 = ax.bar(x + width/2, re_results, width, label='Regulatory Element')

    ax.set_ylabel('Cosine Similarity')
    ax.set_title('Cosine Similarity by Organ')
    ax.set_xticks(x)
    ax.set_xticklabels(geneORGANizer['bodyparts'].keys())
    ax.legend()

    fig.tight_layout()

    plt.savefig('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/vec/new_vwc_each_organs.png', dpi=500)


    # cnt_gene = 0
    # cnt_nogene = 0
    # cnt_re = 0
    # cnt_nore = 0
    # vec_gene = []
    # vec_re = []
    # for organ in geneORGANizer['bodyparts'].keys(): 
    #     for gene in total_geneset:
    #         if gene in vecs:
    #             cnt_gene += 1
    #             vec_gene.append(np.array(cosine_similarity([vecs[gene], vecs[organ]])[0][1]))
    #         else:
    #             cnt_nogene += 1
        
    #     for re in grn['res'].keys():
    #         if re in vecs:
    #             cnt_re += 1
    #             vec_re.append(np.array(cosine_similarity([vecs[re], vecs[organ]])[0][1]))
    #         else:
    #             cnt_nore += 1
    # mean_vec_gene = np.mean(vec_gene, axis=0)
    # median_vec_gene = np.median(vec_gene, axis=0)

    # print(cnt_gene)
    # print(cnt_nogene)
    # print(mean_vec_gene)
    # print(median_vec_gene)

    # mean_vec_re = np.mean(vec_re, axis=0)
    # median_vec_re = np.median(vec_re, axis=0)

    # print(cnt_re)
    # print(cnt_nore)
    # print(mean_vec_re)
    # print(median_vec_re)


    # # 绘制向量列表密度图
    # density1 = gaussian_kde(vec_gene)
    # density2 = gaussian_kde(vec_re)
    # xs = np.linspace(0, 1, 200)
    # plt.plot(xs, density1(xs), label='vec_gene')
    # plt.plot(xs, density2(xs), label='vec_re')

    # # 添加图例和标签
    # plt.title("Vector Density")
    # plt.xlabel("Value")
    # plt.ylabel("Density")
    # plt.legend()
    # plt.savefig('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/vec/new_vec_bet1.png', dpi=500)

    # # 绘制直方图
    # plt.hist(vec_gene, bins=50, alpha=0.5, label='vec_gene')
    # plt.hist(vec_re, bins=50, alpha=0.5, label='vec_re')
    # plt.legend(loc='upper right')
    # plt.show()

    # # 添加图例和标签
    # plt.title("Vector Histogram")
    # plt.xlabel("Value")
    # plt.ylabel("Frequency")
    # plt.legend()
    # plt.savefig('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/vec/new_vec_bet2.png', dpi=500)

def importvec(filename):
    vecs = dict()
    with open(filename, 'r') as f:
        next(f)
        for line in f:
            line = line.split()
            obj = line[0]
            for index in range(1, len(line)):  # organs名字可能是'xxx xxx xxx'形式
                if line[index].isalpha():
                    obj += ' '
                    obj += line[index]
                else:
                    break
            vecs[obj] = [float(x) for x in line[index:]]
    #  测试是否正确拿到organ名称
    # with open('./process_result/lzm_test_vecs.txt', 'w') as f:
    #     f.write(str(vecs))
    print('import vecs done')
    return vecs

def import_pos():
    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    geneORGANizer = datas.import_database_geneORGANizer()
    vecs = importvec('./LiCO/vec/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')

    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))
    
    Genes_under_selection = []
    cnt = 0
    with open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/P.t. troglodytesbottom.txt', 'r') as f:
        next(f)
        next(f)
        for line in f:
            line = line.split('\n')
            gene = line[0].upper()
            # print(gene)
            if gene in total_geneset and vecs:
            # if gene in geneORGANizer['genes']:
                Genes_under_selection.append(gene)
                cnt += 1
                # print(gene)
    print('cnt_match_gene: ', cnt)
    return Genes_under_selection

# 导入与geneORGANizer/hGraph重合的高原适应症受选择G
def import_altitude_genes(opt):
    geneORGANizer = datas.import_database_geneORGANizer()
    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    Genes_under_selection = []
    # if opt == 'all':
    cnt_all = 0
    with open('./LiCO/data/Indications_for_high_altitude_genes_all.txt', 'r') as f:
        print('open _all txt')
        for line in f:
            line = line.split('\n')
            gene = line[0].upper()
            # if gene in geneORGANizer['genes']:
            if gene in total_geneset:
                Genes_under_selection.append(gene)
                cnt_all += 1
    print('cnt_all:', cnt_all)
    # if opt == 'more':
    #     cnt_more = 0
    #     with open('./LiCO/data/Indications_for_high_altitude_genes_more.txt', 'r') as f:  # 不包含只有1次文献验证的G
    #         print('open _more txt')
    #         for line in f:
    #             line = line.split('\n')
    #             gene = line[0].upper()
    #             if gene in geneORGANizer['genes']:
    #                 Genes_under_selection.append(gene)
    #                 cnt_more += 1
    #     print('cnt_more:', cnt_more)

    return Genes_under_selection

# 导入与geneORGANizer/hGraph重合的回声定位受选择G
def import_echo_genes():
    geneORGANizer = datas.import_database_geneORGANizer()
    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    Genes_under_selection = []
    cnt = 0
    with open('./LiCO/echo/gene.txt', 'r') as f:
        for line in f:
            line = line.split('\n')
            gene = line[0].upper()
            # print(gene)
            if gene in total_geneset:
            # if gene in geneORGANizer['genes']:
                Genes_under_selection.append(gene)
                cnt += 1
    print('cnt_echo_gene: ', cnt)

    return Genes_under_selection
# 导入与糖尿病有关的RE
def import_diabetes_res():

    res_under_selection = []
    cnt = 0
    with open('./LiCO/ad/ad_GRN_re_5e4.txt', 'r') as f:
        for line in f:
            line = line.split()
            re = line[0]
            if re not in res_under_selection:
                res_under_selection.append(re)
                cnt += 1
    print('cnt_ad_re', cnt)

    return res_under_selection

# 导入与SpecVar有关的RE 
def import_GWAS_res():

    res_under_selection = []
    cnt = 0
    with open('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/SpecVar/LDL/LDL2RE1e-11_window1_overlap.txt', 'r') as f:
        for line in f:
            line = line.split()
            re = line[0]
            if re not in res_under_selection:
                res_under_selection.append(re)
                cnt += 1
    print('cnt_SpecVar_re', cnt)

    return res_under_selection

# 写入与SpecVar有关的G(由RE转化而来) 
def import_GWAS_genes():
    # grn = datas.import_database_grnhg()
    # res_under_selection = []
    # genes_under_selection = []
    # cnt = 0
    # w = open('./LiCO/SpecVar/EA/EA2RE2G1e-20_window1_overlap.txt', 'w')
    # with open('./LiCO/SpecVar/EA/EA2RE1e-20_window1_overlap.txt', 'r') as f:
    #     for line in f:
    #         line = line.split()
    #         re = line[0]
    #         if re not in res_under_selection:
    #             res_under_selection.append(re)
    #             cnt += 1
    # print('cnt_SpecVar_re', cnt)
    # f.close()
    # cnt = 0
    # for re in res_under_selection:
    #     for gene in grn['res'][re]['reg']:
    #         if gene not in genes_under_selection:
    #             genes_under_selection.append(gene)
    #             w.write(gene)
    #             w.write('\n')
    #             cnt += 1
    # print('cnt_SpecVar_gene', cnt)

    Genes_under_selection = []
    cnt = 0
    with open('./LiCO/SpecVar/EA/EA2RE2G1e-20_window1_overlap.txt', 'r') as f:
        for line in f:
            line = line.split('\n')
            gene = line[0]
            if gene not in Genes_under_selection:
                Genes_under_selection.append(gene)
                cnt += 1
    print('cnt_SpecVar_gene', cnt)
    # print(Genes_under_selection)
    print(len(Genes_under_selection))

    return Genes_under_selection

# 导入与geneORGANizer/hGraph重合的糖尿病有关的G
def import_diabetes_gene():
    geneORGANizer = datas.import_database_geneORGANizer()
    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    Genes_under_selection = []
    cnt = 0
    with open('./LiCO/diabetes_mellitus/diabetes_GRN_listRE2G_5e6_fullMatch.txt', 'r') as f:
        for line in f:
            line = line.split('\n')
            gene = line[0].upper()
            # print(gene)
            if gene in total_geneset:
                Genes_under_selection.append(gene)
                cnt += 1
    print('cnt_echo_gene: ', cnt)

    return Genes_under_selection

# 导入与geneORGANizer/hGraph重合的糖尿病有关的G
def import_T2D_diabetes_gene():
    # cnt = 0
    # grn = datas.import_database_grnhg()
    # w = open('./LiCO/diabetes_mellitus/T2D/SNP_5e-8.txt', 'w')
    # filename = './LiCO/diabetes_mellitus/T2D/associations.csv'
    # with open(filename, 'r') as f:
    #     next(f)
    #     for line in f:
    #         line = line.split(',')
            
    #         raw_re = line[0] # "10:114754088:T:C"
    #         raw_re = eval(raw_re)
    #         raw_re = raw_re.split(':')
    #         chr = raw_re[0]
            
    #         pos = raw_re[1]
            
            
    #         p_val = line[10]
    #         #
    #         #
    #         # nearest_gene = line[26]
    #         # nearest_gene = nearest_gene.replace('[','').replace(']','')
    #         # nearest_gene = eval(nearest_gene)
    #         #
    #         if float(p_val) <= 0.00000005:
    #             cnt += 1
    #             w.write('chr' + str(chr) + '_' + str(pos))
    #             w.write('\n')
    #         # print(chr)
    #         # print(pos)
    #         # print(p_val)
    #         # break

    #     print(cnt)
    # f.close()
    # w.close()

    grn = datas.import_database_grnhg()
    cnt_num = 0
    cnt = 0
    w = open('./LiCO/diabetes_mellitus/T2D/SNP2RE_5e-8.txt', 'w')
    with open('./LiCO/diabetes_mellitus/T2D/SNP_5e-8.txt', 'r') as f:
        for line in f:
            line = line.split('_')
            # raw_re = line[0]
            # print(raw_re)
            # raw_re = raw_re.split(':')
            raw_num = line[0]
            # print(raw_num)
            raw_pos = line[1]
            # print(raw_pos)

            for re_write in grn['res'].keys():
                re = re_write
                re = re.split('_')
                re_num = re[0]
                re_start = re[1]
                re_end = re[2]

                if re_num == raw_num:
                    cnt_num += 1
                    if raw_pos <= re_end and raw_pos >= re_start:
                        cnt += 1
                        w.write(str(re_write))
                        w.write('\n')
        print(cnt)
        print(cnt_num)



    # # 从文件中读取G
    # cnt = 0
    # filename = './LiCO/diabetes_mellitus/T2D/gene_table.csv'
    # w = open('./LiCO/diabetes_mellitus/T2D/re_5e9.txt', 'w')
    # with open(filename, 'r') as f:
    #     next(f)
    #     for line in f:
    
    #         line = line.split(',')
    #         # print(len(line))
    
    #         gene = eval(line[1])
    #         num = line[2].strip('\"')
    #         start = line[3]
    #         end = line[4]    
    
    #         p_val = line[6]
    
    #         if float(p_val) <= 0.000000005:
    #             w.write('chr' + str(num) + '_' + str(start) + '_' + str(end))
    #             w.write('\n')
    #             cnt += 1
    #     print(cnt)
    # f.close()
    # w.close()

    # # 写入受选择gene
    # cnt = 0
    # filename = './LiCO/diabetes_mellitus/T2D/gene_table.csv'
    # w = open('./LiCO/diabetes_mellitus/T2D/gene_5e9.txt', 'w')
    # with open(filename, 'r') as f:
    #     next(f)
    #     for line in f:
    #
    #         line = line.split(',')
    #         # print(len(line))
    #
    #         gene = eval(line[1])
    #
    #
    #         p_val = line[6]
    #
    #         if float(p_val) <= 0.000000005:
    #             w.write(gene)
    #             w.write('\n')
    #             cnt += 1
    #     print(cnt)
    # f.close()
    # w.close()

    # # 返回受选择基因
    # geneORGANizer = datas.import_database_geneORGANizer()
    # grn = datas.import_database_grnhg()
    # ppi = datas.importppi()
    # total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    # Genes_under_selection = []
    # cnt = 0
    # with open('./LiCO/diabetes_mellitus/T2D/gene_5e8.txt', 'r') as f:
    #     for line in f:
    #         line = line.split()
    #         gene = line[0]
    #         if gene in geneORGANizer['genes']:
    #         # if gene in total_geneset:
    #             if gene not in Genes_under_selection:
    #                 Genes_under_selection.append(gene)
    #                 cnt += 1
    # print('cnt_T2D : ', cnt)
    # f.close()
    # print(Genes_under_selection)
    # print(len(Genes_under_selection))
    
    # return Genes_under_selection

# 导入SpecVar中性状相关GWAS数据
def import_GWAS():
    # cnt = 0
    # # w_LDL = open('./LiCO/SpecVar/LDL_1e-11.txt', 'w')
    # # w_TC = open('./LiCO/SpecVar/TC_1e-11.txt', 'w')
    # w_EA = open('./LiCO/SpecVar/EA/EA_1e-20.txt', 'w') # 314

    # # filename = './LiCO/SpecVar/Mc_TC.txt'
    # # filename = './LiCO/SpecVar/GWAS_CP_all.txt'
    # # filename = './LiCO/SpecVar/Mc_TC.txt'
    # # SNP_hg18	SNP_hg19	rsid	A1	A2	beta	se	N	P-value	Freq.A1.1000G.EUR

    # # 读入tsv文件
    # filename = './LiCO/SpecVar/EA/EA.tsv'
    # cnt = 0
    # with open(filename) as f:
    #     next(f)
    #     tsvreader = csv.reader(f, delimiter='\t')
    #     for line in tsvreader:
    #         # print(line)

    #         chr = line[11]
    #         # print(chr)
    #         pos = line[12]
    #         # print(pos)

    #         p_val = line[27]
    #         # print(p_val)
    #         if float(p_val) <= 0.00000000000000000001:
    #             cnt += 1
    #             w_EA.write('chr' + chr + ' ' + pos)
    #             w_EA.write('\n')
    #         # break
    #     print(cnt)
    # f.close()
    # w_EA.close()


    # with open(filename, 'r') as f:
    #     next(f)
    #     for line in f:
    
    #         line = line.split()
    #         # re = line[1]
    #         chr = line[11]
    #         print(chr)
    #         pos = line[12]
    #         print(pos)
    
    
    #         p_val = line[27]
    #         print(p_val)
    #         break
    
    #         if float(p_val) <= 0.00000000001:
    #             cnt += 1
    #             w_CP.write('chr' + chr + ' ' + pos)
    #             w_CP.write('\n')

    #             # w_TC.write(re)
    #             # w_TC.write('\n')
    
    #             # w_TC.write(re_hg19)
    #             # w_TC.write('\n')
    #     print(cnt)
    # f.close()

    cnt = 0
    grn = datas.import_database_grnhg()

    # w_TC = open('./LiCO/SpecVar/TC2RE1e-11_window1_overlap.txt', 'w')
    # w_CP = open('./LiCO/SpecVar/CP2RE1e-11_window1_overlap.txt', 'w')
    w_EA = open('./LiCO/SpecVar/EA/EA2RE1e-20_window1_overlap.txt', 'w') # 135
    # filename = './LiCO/SpecVar/LDL_0.01.txt'
    # filename = './LiCO/SpecVar/TC_1e-11.txt'
    filename = './LiCO/SpecVar/EA/EA_1e-20.txt'
    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            # raw_re = line[0]
            # print(raw_re)
            # raw_re = raw_re.split(':')
            raw_num = line[0]
            # print(raw_num)
            raw_pos = line[1]
            # print(raw_pos)
            # break


            for re_write in grn['res'].keys():
                re = re_write
                re = re.split('_')
                re_num = re[0]
                re_start = re[1]
                re_end = re[2]

                # print(re_num)
                # print(re_start) 
                # print(re_end)

                # 转化为GRN中的RE的2种形式：严格注释、窗口注释
                # 严格注释
                if re_num == raw_num:
                    if raw_pos <= re_end and raw_pos >= re_start:
                        cnt += 1
                        # w_TC.write(str(re_write))
                        # w_TC.write('\n')
                        w_EA.write(str(re_write))
                        w_EA.write('\n')
                # else:
                #     print(re_num) 
                #     print(raw_num)
  
    print(cnt)
    f.close()
    w_EA.close()

# 导入与阿尔兹海默症有关的RE
def import_ad_res():

    res_under_selection = []
    cnt = 0
    with open('./LiCO/diabetes_mellitus/diabetes_GRN_re_5e5_fullMatch.txt', 'r') as f:
        for line in f:
            line = line.split()
            re = line[0]
            if re not in res_under_selection:
                res_under_selection.append(re)
                cnt += 1
    print('cnt_ad_re', cnt)

    return res_under_selection

# 以TopN的方式 为G注释B
def generateDict_gene_predict_bodypart(topN):
    vecs = importvec('./LiCO/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')
    geneORGANizer = datas.import_database_geneORGANizer()
    ppi = datas.importppi()

    prob = dict()
    # 以gene预测TopN bodypart
    gene_predict_bodypart_dict = dict()

    # Genes_under_selection = import_altitude_genes(opt='all')  # 高原适应症G
    Genes_under_selection = import_echo_genes()  # 回声定位G

    M = len(ppi.keys())
    count = 0
    for gene in Genes_under_selection:
        if gene in vecs:
            if gene not in gene_predict_bodypart_dict:
                gene_predict_bodypart_dict[gene] = set()
            prob[gene] = OrderedDict()
            for bodypart in geneORGANizer['bodyparts'].keys():
                prob[gene][bodypart] = cosine_similarity([vecs[gene], vecs[bodypart]])[0][1]

            vd = OrderedDict(sorted(prob[gene].items(), key=lambda t: t[1], reverse=True))

            i = 0
            for k in vd.keys():
                gene_predict_bodypart_dict[gene].add(k)
                i += 1
                if i == topN:
                    break

            count = count + 1
            if count % 10 == 0:
                print(count, ' / ', M)

    # print(gene_predict_bodypart_dict)
    return gene_predict_bodypart_dict

# 取得B的阈值
def getthres():
    fmax = dict()
    bodyparts = set()
    tfs = set()
    # re-g-b-threshold.txt
    with open('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/threshold/' + 'gO_B_G_threshold.txt', 'r') as f:  # 由文件'calthres.py'中的方法cal()得到，其中是term和对应的threshold
        next(f)  # 第一行是标题
        for line in f:
            line = line.split()  # 格式：gene  threshigh  thresfmax  threslow
            # 因为组织名字有可能为 xxx xxx xxx 故需要同importvecs()一样处理
            obj = line[0]

            for index in range(1, len(line)):  # organs名字可能是'xxx xxx xxx'形式
                if line[index].isalpha():
                    obj += ' '
                    obj += line[index]
                else:
                    break
            fmax[obj] = float(line[index + 1])  # 得到[term——thres]
            bodyparts.add(obj)  # 得到一系列具有阈值threshold的term

    return fmax, bodyparts

# 高原适应症
def altitude():
    Genes_under_selection = import_altitude_genes(opt='all')  # 所有文献记载过的G
    # Genes_under_selection = import_altitude_genes(opt='more')  # 文献中出现超过1次的G
    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}

    vecs = importvec('./LiCO/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')
    grn = datas.import_database_grnhg()
    ppi = datas.importppi()

    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    # 查询异质网络和高原适应症G的交集情况
    # cnt_1 = 0
    # cnt_2 = 0
    # cnt_overlap = 0
    # for gene_1 in total_geneset:
    #     cnt_1 += 1
    #     for gene_2 in Genes_under_selection:
    #         cnt_2 += 1
    #         if gene_1.upper() == gene_2.upper():
    #             cnt_overlap += 1
    # print(len(Genes_under_selection))
    # print(cnt_overlap)
    # 为150个交集G注释B
    # 选用2种方法：topN、取fmax时的阈值

    # topN
    # 统计频次
    b_fre = dict()
    gene_predict_topN_b = generateDict_gene_predict_bodypart(topN=3)

    with open('./LiCO/altitude/Allegene_predict_top3_b.txt', 'w') as w:
        with open('./LiCO/altitude/Allgene_predict_top3_b_fre.txt', 'w') as w_fre:
            for gene in gene_predict_topN_b:

                w.write(gene)
                w.write('\t')
                for b in gene_predict_topN_b[gene]:
                    w.write(b)
                    w.write('\t')
                    if b not in b_fre:
                        b_fre[b] = 0
                    b_fre[b] += 1

                w.write('\n')


            # 输出b_fre
            # 字典按值排序
            b_order = sorted(b_fre.items(), key=lambda x: x[1], reverse=True)
            for b in b_order:  # ('intestine', 1)
                w_fre.write(b[0])
                w_fre.write('\t')



                w_fre.write(str(b[1]))
                w_fre.write('\n')
        w_fre.close()
    w.close()


    # # 取fmax时的阈值
    # fmax, bodyparts = getthres()
    # prob = dict()
    # count = 0
    # b_fre = dict()
    # with open('./LiCO/altitude/' + 'MoreGene_predict_fmax_b.txt', 'w') as w:
    #     with open('./LiCO/altitude/MoreGene_predict_fmax_b_fre.txt', 'w') as w_fre:
    #         for gene in Genes_under_selection:
    #             if gene in vecs:
    #                 if gene not in prob:
    #                     prob[gene] = dict()
    #             else:
    #                 continue
    #             for bodypart in bodyparts:
    #                 if bodypart in vecs:
    #                     prob[gene][bodypart] = cosine_similarity([vecs[gene], vecs[bodypart]])[0][1]
    #             w.write(gene + '\t')
    #
    #             for bodypart in bodyparts:
    #                 if bodypart in vecs:
    #                     if prob[gene][bodypart] > fmax[bodypart]:
    #                         w.write(bodypart + '\t')
    #                         if bodypart not in b_fre:
    #                             b_fre[bodypart] = 0
    #                         b_fre[bodypart] += 1
    #
    #             w.write('\n')
    #         # 输出b_fre
    #         # 字典按值排序
    #         b_order = sorted(b_fre.items(), key=lambda x: x[1], reverse=True)
    #         for b in b_order:
    #             w_fre.write(b[0])
    #             w_fre.write('\t')
    #
    #             w_fre.write(str(b[1]))
    #             w_fre.write('\n')
    #
    #     w_fre.close()
    # w.close()

# 回声定位
def echo():
    Genes_under_selection = import_echo_genes()
    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}

    vecs = importvec('./LiCO/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')
    grn = datas.import_database_grnhg()
    ppi = datas.importppi()

    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    # 查询异质网络和echoG的交集情况
    # cnt_1 = 0
    # cnt_2 = 0
    # cnt_overlap = 0
    # for gene_1 in total_geneset:
    #     cnt_1 += 1
    #     for gene_2 in Genes_under_selection:
    #         cnt_2 += 1
    #         if gene_1.upper() == gene_2.upper():
    #             cnt_overlap += 1
    # print(len(Genes_under_selection))
    # print(cnt_overlap)
    # 为108个交集G注释B
    # 选用2种方法：topN、取fmax时的阈值

    # topN
    # 统计频次
    b_fre = dict()
    gene_predict_topN_b = generateDict_gene_predict_bodypart(topN=10)

    # with open('./LiCO/echo/gene_predict_top10_b.txt', 'w') as w:
    #     with open('./LiCO/echo/gene_predict_top10_b_fre.txt', 'w') as w_fre:
    #         for gene in gene_predict_topN_b:
    #             w.write(gene)
    #             w.write('\t')
    #             for b in gene_predict_topN_b[gene]:
    #                 w.write(b)
    #                 w.write('\t')
    #                 if b not in b_fre:
    #                     b_fre[b] = 0
    #                 b_fre[b] += 1
    #
    #             w.write('\n')
    #
    #         # 输出b_fre
    #         # 字典按值排序
    #         b_order = sorted(b_fre.items(), key=lambda x: x[1], reverse=True)
    #         for b in b_order:  # ('intestine', 1)
    #             w_fre.write(b[0])
    #             w_fre.write('\t')
    #
    #             w_fre.write(str(b[1]))
    #             w_fre.write('\n')
    #     w_fre.close()
    # w.close()
    # 做富集分析：超几何分布

# geneORGANizer中每个B对G的覆盖情况
def gO_B_link_G():
    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}
    dict_B_link_G = dict()

    with open('./LiCO/altitude/geneORGANizer_b_link_g.txt', 'w') as w:
        for b in geneORGANizer['bodyparts'].keys():
            if b not in dict_B_link_G:
                dict_B_link_G[b] = 0
            dict_B_link_G[b] += len(geneORGANizer['bodyparts'][b])

        b_order = sorted(dict_B_link_G.items(), key=lambda x: x[1], reverse=True)
        for b in b_order:  # ('intestine', 1)
            w.write(b[0])
            w.write('\t')

            w.write(str(b[1]))
            w.write('\n')

    w.close()

# organs间关系
# 根据一定阈值，建立B之间的联系
def link_b():

    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}
    vecs = importvec('./LiCO/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')

    prob = dict()
    record = dict()
    # 根据fmax阈值建立联系
    fmax, bodyparts = getthres()

    cossim_dict = dict()
    with open('./LiCO/organs_link/newForm1_ogansLink_fmax.txt', 'w') as w:
        for bodypartA in geneORGANizer['bodyparts']:
            # w.write(bodypartA)
            # w.write('\t')
            cossim_dict[bodypartA] = set()
            prob[bodypartA] = dict()
            record[bodypartA] = set()
            for bodypartB in geneORGANizer['bodyparts']:
                if bodypartB != bodypartA:
                    record[bodypartB] = set()
                    prob[bodypartA][bodypartB] = cosine_similarity([vecs[bodypartA], vecs[bodypartB]])[0][1]

                else:
                    continue
                if prob[bodypartA][bodypartB] >= fmax[bodypartA]:
                    # 如果关系对没有出现过
                    if bodypartB not in record[bodypartA] or bodypartA not in record[bodypartB]:
                        w.write(bodypartA)
                        w.write('\t')
                        w.write(bodypartB)
                        w.write('\n')

                        record[bodypartA].add(bodypartA)
                        record[bodypartB].add(bodypartB)

                    cossim_dict[bodypartA].add(bodypartB)
            w.write('\n')
    w.close()

    # # 基于geneORGANizer的“Jaccard相似度”，设定阈值连边
    # prob = dict()
    #
    # sum = list()
    # cnt = 0
    # jaccardsim_dict = dict()
    # w = open('./LiCO/organs_link/ogansLink_jaccard_fmax.txt', 'w')
    # for organA in geneORGANizer['bodyparts']:
    #     if organA not in prob:
    #         prob[organA] = dict()
    #         w.write(organA)
    #         jaccardsim_dict[organA] = set()
    #     for organB in geneORGANizer['bodyparts']:
    #         if organB != organA:
    #             # 计算交集和并集
    #             listA = set()
    #             listB = set()
    #             for gene in geneORGANizer['genes']:
    #                 if gene in geneORGANizer['bodyparts'][organA]:
    #                     listA.add(gene)
    #                 if gene in geneORGANizer['bodyparts'][organB]:
    #                     listB.add(gene)
    #             # 交集
    #             list_overlap = listA & listB
    #             len_o = len(list_overlap)
    #             # 并集
    #             list_union = listA | listB
    #             len_u = len(list_union)
    #
    #
    #
    #
    #             prob[organA][organB] = abs(len_o)/abs(len_u)
    #             # sum += prob[organA][organB]
    #             # sum.append(prob[organA][organB])
    #             # cnt += 1
    #
    #             # if prob[organA][organB] >= 0.2331:
    #             # if prob[organA][organB] >= 0.2748:
    #             if prob[organA][organB] >= fmax[organA]:
    #                 w.write('\t')
    #                 w.write(organB)
    #                 jaccardsim_dict[organA].add(organB)
    #     w.write('\n')
    #
    #
    # # 筛选出新的关系对：余弦相似度 VS jaccard相似度
    # w = open('./LiCO/organs_link/overlap_relationship_withFmax.txt', 'w')
    # for key in cossim_dict:
    #     set1 = cossim_dict[key]
    #     set2 = jaccardsim_dict[key]
    #
    #     set3 = set1 & set2
    #
    #     w.write(key)
    #     w.write('\t')
    #     for item in list(set3):
    #         w.write(item)
    #         w.write('\t')
    #     w.write('\n')
    #
    # w.close()

# 预估阈值，寻找与organs无特异性/有特异性的TFs
def findSpecificTFs():
    grn = datas.import_database_grnhg()  # return {'tfs': tfs, 'gene_tfs': gene_tf}
    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}
    vecs = importvec('./LiCO/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')

    prob = dict()
    # 根据fmax阈值建立联系
    fmax, bodyparts = getthres()

    tf_fre_dict = dict()
    with open('./LiCO/SpecificTFs/SpecificTFs_fmax_orderFre.txt', 'w') as w:
        for tf in grn['tfs']:
            tf_fre_dict[tf] = 0
            cnt = 0
            # w.write(tf)
            # w.write('\t')
            prob[tf] = dict()
            for bodypart in geneORGANizer['bodyparts']:
                prob[tf][bodypart] = cosine_similarity([vecs[bodypart], vecs[tf]])[0][1]

                if prob[tf][bodypart] > fmax[bodypart]:
                    # w.write(bodypart)
                    # w.write('\t')
                    cnt += 1
                    tf_fre_dict[tf] += 1
            # w.write(str(cnt))
            # w.write('\n')

        # 排序 tf_fre_dict，再打印
        tf_fre_order = sorted(tf_fre_dict.items(), key=lambda x: x[1], reverse=True)
        for tf in tf_fre_order:
            w.write(tf[0])
            w.write('\t')

            w.write(str(tf[1]))
            w.write('\n')

    w.close()

def enrichment():
    '''
    首先基于geneORGANizer得到一批富集的organs；
	然后基于异质网络得到一批富集的organs；
	    背景基因：整个异质网络的基因
		背景连边情况：基于阈值

		受选择基因：文件
		受选择基因的连边情况：基于阈值
    '''
    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}
    # print('len of gO gene:', len(geneORGANizer['genes']))

    # Genes_under_selection = import_echo_genes()
    # Genes_under_selection = import_altitude_genes(all)
    # Genes_under_selection = import_diabetes_gene()
    # Genes_under_selection = import_T2D_diabetes_gene()
    # Genes_under_selection = import_GWAS_genes()
    # Genes_under_selection = import_GWAS_genes()
    Genes_under_selection = import_pos()
    print('len of sel gene:', len(Genes_under_selection))


    grn = datas.import_database_grnhg()
    ppi = datas.importppi()

    vecs = importvec('./LiCO/vec/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')
    # vecs = importvec('./LiCO/' + '10_15_7path_addGB140w_epoch100_2_4layerTrain_humanGRN_2_vec.txt')


    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))
    print('len of hGraph gene:', len(total_geneset))

    total_vec_geneset = set()
    cnt_Gvec = 0

    for gene in total_geneset:
        if gene in vecs:
            total_vec_geneset.add(gene)
            cnt_Gvec += 1
    print('len(total_vec_geneset) : ', len(total_vec_geneset))

    organ_p = dict()
    # w_s = open('./LiCO/echo/V4/P_gO_enrichment_30Genes.txt', 'w')
    w_cdf = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/gO_troglodytesbottom.txt', 'w')
    # w_sf = open('./LiCO/echo/V4/Psf_gO_enrichment_30Genes.txt', 'w')
    # 基于geneORGANize得到一批富集organs
    for organ in geneORGANizer['bodyparts']:
        if organ not in organ_p:
            organ_p[organ] = 0
        # w_s.write(organ + '\t')
        w_cdf.write(organ + '\t')
        # w_sf.write(organ + '\t')
        print('current organ: ', organ)
        # 当前organ在geneORGANizer中的注释数量
        cnt_gO = len(geneORGANizer['bodyparts'][organ])
        print('cnt_gO: ', cnt_gO)
        # 当前organ在受选择G中的注释数量
        cnt_slG = 0
        for gene in Genes_under_selection:
            if gene in geneORGANizer['bodyparts'][organ]:
                cnt_slG += 1
        print('cnt_slG: ', cnt_slG)
        # 对当前organs进行超几何分布检验
        # 抽样10w次，每次抽XX个G，该organs注释次数>=cnt_slG的概率
        # s = np.random.hypergeometric(cnt_gO, len(geneORGANizer['genes'])-cnt_gO, len(Genes_under_selection), size=100000)
        # p = sum(s >= cnt_slG)/100000
        # organ_p[organ] = p
        # w_s.write(str(p) + '\n')
        # print(p)
    
        # 更换为另一种超几何分布形式：
        # 小集合被选中-1、背景、背景被选中、小集合
        # pv_sf = stats.hypergeom.sf(cnt_slG-1, len(geneORGANizer['genes']), cnt_gO, len(Genes_under_selection))
        pv_cdf = 1 - stats.hypergeom.cdf(cnt_slG-1, len(geneORGANizer['genes']), cnt_gO, len(Genes_under_selection))
        organ_p[organ] = pv_cdf
        w_cdf.write(str(pv_cdf) + '\n')
        # w_sf.write(str(pv_sf) + '\n')
    
        # print('p : ', p)
        # print('pv_sf : ', pv_sf)
        # print('pv_cdf : ', pv_cdf)





    # 基于hGraph得到一批富集organs
    fmax, bodyparts = getthres()

    organ_p = dict()
    # w_s = open('./LiCO/SpecVar/P_hGraph_enrichment_458Genes.txt', 'w')
    w_cdf = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/nO_troglodytesbottom.txt', 'w')
    # w_sf = open('./LiCO/SpecVar/Psf_hGraph_enrichments.txt', 'w')
    for organ in geneORGANizer['bodyparts']:
        if organ not in organ_p:
            organ_p[organ] = 0
        # w_s.write(organ + '\t')
        w_cdf.write(organ + '\t')
        # w_sf.write(organ + '\t')

        print('curren organ : ', organ)
        # 当前organ在hGraph中的注释情况
        cnt_hG = 0
        cnt_slG = 0
        for gene in total_geneset:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_hG += 1
            else:
                continue

        for gene in Genes_under_selection:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_slG += 1
            else:
                continue
        print('cnt_hG : ', cnt_hG)
        print('cnt_slG: ', cnt_slG)

        # pv_sf = stats.hypergeom.sf(cnt_slG - 1, len(total_vec_geneset), cnt_hG, len(Genes_under_selection))
        pv_cdf = 1 - stats.hypergeom.cdf(cnt_slG - 1, len(geneORGANizer['genes']), cnt_hG, len(Genes_under_selection))
        # pv_sf = stats.hypergeom.sf(cnt_slG - 1, len(geneORGANizer['genes']), cnt_hG, len(Genes_under_selection))
        # pv_cdf = 1 - stats.hypergeom.cdf(cnt_slG - 1, len(total_vec_geneset), cnt_hG, len(Genes_under_selection))

        # organ_p[organ] = pv_cdf
        w_cdf.write(str(pv_cdf) + '\n')
        # w_sf.write(str(pv_sf) + '\n')

        # print(p)

# RE在organs上的富集
def enrichment_re():
    '''
	然后基于异质网络得到一批富集的organs；
	    背景RE：整个异质网络的RE
		背景连边情况：基于阈值

		受选择RE：文件
		受选择RE的连边情况：基于阈值
    '''

    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}

    # res_under_selection = import_diabetes_res()
    res_under_selection = import_GWAS_res()


    grn = datas.import_database_grnhg()


    vecs = importvec('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/vec/10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')
    # 新增加REB之后的模型
    # vecs = importvec('./LiCO/' + '10_15_7path_addGB140w_epoch50_2_4layerTrain_humanGRN_2_vec.txt')
    # # vecs = importvec('./LiCO/' + '10_15_7path_addGB140w_epoch100_2_4layerTrain_humanGRN_2_vec.txt')
    # vecs = importvec('./LiCO/' + '10_15_7path_addGB140w_epoch50_4layerTrain_humanGRN_2_vec.txt')
    # vecs = importvec('./LiCO/' + '10_15_7path_addGB140w_epoch100_4layerTrain_humanGRN_2_vec.txt')

    cnt_REvec = 0
    for re in grn['res'].keys():
        if re in vecs:
            cnt_REvec += 1
    print('cnt_REvec : ', cnt_REvec)

    # organ_p = dict()
    # w = open('./LiCO/SpecVar/LDL_enrichment.txt', 'w')
    # # 基于geneORGANize得到一批富集organs
    # for organ in geneORGANizer['bodyparts']:
    #     if organ not in organ_p:
    #         organ_p[organ] = 0
    #     w.write(organ + '\t')
    #     # print('current organ: ', organ)
    #     # 当前organ在geneORGANizer中的注释数量
    #     cnt_gO = len(geneORGANizer['bodyparts'][organ])
    #     # print('cnt_gO: ', cnt_gO)
    #     # 当前organ在受选择G中的注释数量
    #     cnt_slG = 0
    #     for gene in Genes_under_selection:
    #         if gene in geneORGANizer['bodyparts'][organ]:
    #             cnt_slG += 1
    #     # print('cnt_slG: ', cnt_slG)
    #     # 对当前organs进行超几何分布检验
    #     # 抽样10w次，每次抽30个G，该organs注释次数>=cnt_slG的概率
    #     s = np.random.hypergeometric(cnt_gO, len(geneORGANizer['genes'])-cnt_gO, len(Genes_under_selection), 100000)
    #     p = sum(s >= cnt_slG)/100000
    #     organ_p[organ] = p
    #     w.write(str(p) + '\n')
    #     # print(p)


    # 基于hGraph得到一批富集organs
    fmax, bodyparts = getthres()

    organ_p = dict()
    w = open('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/SpecVar/LDL/(own)LDL_hGraph_enrichment_1e-11.txt', 'w')
    enrichment_results = []
    for organ in geneORGANizer['bodyparts']:
        if organ not in organ_p:
            organ_p[organ] = 0
        w.write(organ + '\t')
        print('curren organ : ', organ)
        # 当前organ在hGraph中的注释情况
        cnt_hG = 0
        cnt_slRE = 0
        for re in grn['res']:
            if re in vecs:
                if cosine_similarity([vecs[re], vecs[organ]])[0][1] >= fmax[organ] * (22/27):
                    cnt_hG += 1
            else:
                continue
        
        overlap_function_res = []
        for re in res_under_selection:
            if re in vecs:
                if cosine_similarity([vecs[re], vecs[organ]])[0][1] >= fmax[organ] * (22/27):
                    cnt_slRE += 1
                    overlap_function_res.append(re)
            else:
                continue
        print('cnt_hG : ', cnt_hG)
        print('cnt_slRE: ', cnt_slRE)

        # 对当前organs进行超几何分布检验
        # 抽样10w次，每次抽len(res_under_selection)个G，该organs注释次数>=cnt_slG的概率
        # s = np.random.hypergeometric(cnt_hG, cnt_R
        # Evec-cnt_hG, len(res_under_selection), 100000)
        # p = sum(s >= cnt_slRE)/100000
        # organ_p[organ] = p
        # w.write(str(p) + '\n')

        M = cnt_REvec
        n = cnt_hG
        N = len(res_under_selection)
        k = cnt_slRE

        print(f"M:{M}, n:{n}, N:{N}, k:{k}")

        p_value = hypergeom.sf(k - 1, M, n, N)
        enrichment_results.append({'organs': organ, 'p_value': p_value, 'res': overlap_function_res})

        pv_cdf = 1 - stats.hypergeom.cdf(cnt_slRE - 1, cnt_REvec, cnt_hG, len(res_under_selection))
        organ_p[organ] = pv_cdf
        w.write(str(pv_cdf) + '\n')
        # print(p)

    # 对p值进行多重检验校正
    p_values = [r['p_value'] for r in enrichment_results]
    # rejected, p_values_corrected, alpha, _ = multipletests(p_values, method='fdr_bh', alpha=0.05)
    p_values_corrected = fdrcorrection(p_values)[1]

    # 将校正后的p值添加到结果中
    for i, result in enumerate(enrichment_results):
        result['p_value_corrected'] = p_values_corrected[i]
        result['rejected'] = p_values_corrected[i] <= 0.05

    # 根据校正后的p值对结果进行排序
    enrichment_results = pd.DataFrame(enrichment_results).sort_values('p_value_corrected')

    # 输出结果
    print(enrichment_results)
    enrichment_results.to_csv('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/SpecVar/LDL/LDL_hGraph_enrichment_1e-11.csv', index=False)

def FDR():
    myMap = dict()
    with open('./LiCO/diabetes_mellitus/T2D/Pcdf_hGraph_enrichment_103Genes.txt') as f:
        for line in f:
            line = line.split()
            organ = line[0]
            for index in range(1, len(line)):
                if line[index].isalpha():
                    organ += ' '
                    organ += line[index]
                else:
                    break
            p_val = line[index]

            if organ not in myMap:
                myMap[organ] = p_val
                # print('organ : ', organ)
                # print('p_val : ', p_val)



    w = open('./LiCO/diabetes_mellitus/T2D/Pcdf_hGraph_enrichment_103Genes_FDR_001.txt', 'w')
    c = 0
    plist = [float(p_val) for p_val in myMap.values()]
    length = len(plist)
    psort = [float(i) for i in plist]
    psort.sort()
    psort.reverse()
    for key, value in myMap.items():
        q = float(value) * length/(psort.index(float(value)) + 1)
        ## 没有执行 FDR(i) = min{FDR(i), FDR(i+1)}, 得到的结果应该更严格
        if q <= 0.01:
            w.write(key + '\t' + value + '\t' + str(q) + '\n')
            c += 1
    w.close()

def fdr_calculate(file_path, q_value=0.01, save_path='FDR_results.txt'):
    my_map = dict()
    with open(file_path) as f:
        for line in f:
            line = line.split()
            organ = line[0]
            for index in range(1, len(line)):
                if line[index].isalpha():
                    organ += ' '
                    organ += line[index]
                else:
                    break
            p_val = line[index]

            if organ not in my_map:
                my_map[organ] = p_val

    psort = sorted(map(float, my_map.values()), reverse=True)
    indices = [psort.index(float(value)) + 1 for value in my_map.values()]

    with open(save_path, 'w') as w:
        c = 0
        for key, value in my_map.items():
            rank = indices[c]
            fdr = float(value) * len(indices) / rank
            if fdr <= q_value:
                w.write(f'{key}\t{value}\t{fdr:.3f}\n')
                c += 1

    print(f'Finished! FDR results are saved in {save_path}.')


def computeFDR(parray, q=0.05):
    if min(parray) < 0 or max(parray) > 1: raise ValueError("请检查P值是否在0~1范围内！\n")

    probs = sorted(parray)
    l = len(probs)

    correct = sum([1 / i for i in range(1, l + 1)])

    fdr = [i / l * (q / correct) for i in range(1, l + 1)]

    sig = [0] * l

    for i in range(l):
        if probs[i] <= fdr[i]:
            sig[i] = 1

    maxsig = max([i * j for i, j in zip(sig, range(1, l))])

    return 0 if maxsig == 0 else probs[maxsig]

# 读取GWAS糖尿病SNP数据,并转化为RE 12/7改进策略：RE再转化为其调控的G
def diabetes():
    # filename = './LiCO/diabetes_mellitus/GCST90161239_buildGRCh37.tsv'
    # cnt = 0
    # w = open('./LiCO/diabetes_mellitus/diabetes_re_5e5_fullMatch.txt', 'w')
    # with open(filename, 'r') as f:
    #     next(f)
    #     for line in f:
    #         line = line.split()
    #         chr = line[0]
    #         pos = line[1]
    #         p_val = line[3]

    #         if float(p_val) <= 0.00005:
    #             re = 'chr' + chr + '_' + pos
    #             w.write(re)
    #             w.write('\n')
    # f.close()
    # w.close()

    geneORGANizer = datas.import_database_geneORGANizer()  # return {'genes': genes, 'bodyparts': bodyparts}
    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))


    # # 为diabetes_re.txt中的RE形式分配真实RE
    # cnt_fullmatch = 0
    # cnt_unfullmatch = 0
    # cnt_dismatch = 0
    # flag_fullmatch = 0
    # # re_set = set()
    # re_list = list() # 不去重re模型
    # w = open('./LiCO/diabetes_mellitus/T2D/grnRE_5e8.txt', 'w')
    # with open('./LiCO/diabetes_mellitus/T2D/re_5e8.txt', 'r') as f:
    #     for line in f:
    #         line = line.split('_')
    #         # chr = line[0]
    #         # pos = line[1]

    #         chr = line[0]
    #         start = line[1]
    #         end = line[2]

    #         # 遍历RE
    #         min_dis = 1000 # 1kb
    #         nearest_re = str()
    #         flag_fullmatch = 0
    #         for origin_re in grn['res']:
    #             re = origin_re.split('_')
    #             chr_ = re[0]
    #             start_ = re[1]
    #             end_ = re[2]

    #             # 进行比对,染色体序号相等，进行下一步区域比较
    #             if chr_ == chr:
    #                 if int(start) >= int(start_) and int(end) <= int(end_):
    #                     cnt_fullmatch += 1
    #                     re_list.append(origin_re)
    #                     flag_fullmatch = 1
    #                     # w.write(origin_re)
    #                     # w.write('\n')
    #                     # 如果在范围内，则用目前RE注释，不考虑GRN中的其它RE
    #                     # break
    #                 # 不在RE范围内，但满足一定要求的：以最近RE进行注释
    #                 # else:
    #                 #     dis = min(abs(int(pos) - int(start)), abs(int(pos) - int(end)))
    #                 #     if dis < min_dis:
    #                 #         min_dis = dis
    #                 #         nearest_re = origin_re
    #                 #     # cnt_match += 1
    #                 #     # re_set.add(origin_re)
    #                 #     # w.write(origin_re)
    #                 #     # w.write('\n')

    #             # 染色体序号不等，continue，考虑下一个re
    #             else:
    #                 cnt_dismatch += 1
    #                 # 染色体序号不匹配，continue
    #                 continue
    #         # if flag_fullmatch != 1:
    #         #     # 序号匹配，为其注释一个最近RE
    #         #     if min_dis < 1000:
    #         #         cnt_unfullmatch += 1
    #         #         print('current min_dis: ', min_dis)
    #         #         re_set.add(nearest_re)
    #         #         # print(min_dis)

    # print('cnt_fullmatch : ', cnt_fullmatch)
    # print('cnt_unfullmatch : ', cnt_unfullmatch)
    # f.close()
    # for re in re_list:
    #     w.write(re)
    #     w.write('\n')

    # w.close()

    vecs = importvec('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/vec/10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')
    # 将受选择RE转化为受选择G
    # seleted_gene_set = set()
    seleted_gene_set = list()
    w1 = open('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/diabetes_mellitus/T2D/SNP2RE2_TOP1totalG_5e8.txt', 'w')
    w2 = open('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/diabetes_mellitus/T2D/SNP2RE2_TOP1grnG_5e8.txt', 'w')
    with open('/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/LiCO/diabetes_mellitus/T2D/SNP2RE_5e-8.txt', 'r') as f:
        for line in f:
            line = line.split()
            re = line[0]
            max_sim = -1
            max_sim_gene = None
            for gene in total_geneset:
                if gene in vecs:
                    sim = cosine_similarity([vecs[re], vecs[gene]])[0][1]
                    if sim > max_sim:
                        max_sim_gene = gene
            if gene not in seleted_gene_set:
                seleted_gene_set.append(max_sim_gene)
                w1.write(max_sim_gene)
                w1.write('\n')

            max_sim = -1
            max_sim_gene = None
            for gene in grn['gene_tfs']:
                if gene in vecs:
                    sim = cosine_similarity([vecs[re], vecs[gene]])[0][1]
                    if sim > max_sim:
                        max_sim_gene = gene
            if gene not in seleted_gene_set:
                seleted_gene_set.append(max_sim_gene)
                w2.write(max_sim_gene)
                w2.write('\n')
    f.close()
    w1.close()
    w2.close()

# 读取GWAS中Alzheimer's disease的SNP数据，并转化为RE
def ad():
    filename1 = './LiCO/ad/GCST90129599_buildGRCh38.tsv'
    filename2 = './LiCO/ad/GCST90129600_buildGRCh38.tsv'
    cnt = 0
    w = open('./LiCO/ad/ad_re_5e4.txt', 'a')
    with open(filename1, 'r') as f:
        next(f)
        for line in f:
            line = line.split()
            chr = line[0]
            pos = line[1]
            p_val = line[6]

            if float(p_val) <= 0.0005:
                # cnt += 1
                re = 'chr' + chr + '_' + pos
                w.write(re)
                w.write('\n')
    # print(cnt)
    f.close()
    w.close()

    grn = datas.import_database_grnhg()
    # 为diabetes_re.txt中的RE形式分配真实RE
    cnt_fullmatch = 0
    cnt_unfullmatch = 0
    cnt_dismatch = 0
    flag_fullmatch = 0
    re_set = set()
    w = open('./LiCO/ad/ad_GRN_re_5e4.txt', 'a')
    with open('./LiCO/ad/ad_re_5e4.txt', 'r') as f:
        for line in f:
            line = line.split('_')
            chr = line[0]
            pos = line[1]

            # 遍历RE
            min_dis = 1000  # 1kb
            nearest_re = str()
            flag_fullmatch = 0
            for origin_re in grn['res']:
                re = origin_re.split('_')
                chr_ = re[0]
                start = re[1]
                end = re[2]

                # 进行比对,染色体序号相等，进行下一步区域比较
                if chr_ == chr:
                    if int(pos) >= int(start) and int(pos) <= int(end):
                        cnt_fullmatch += 1
                        re_set.add(origin_re)
                        flag_fullmatch = 1
                        # w.write(origin_re)
                        # w.write('\n')
                        # 如果在范围内，则用目前RE注释，不考虑GRN中的其它RE
                        break
                    # 不在RE范围内，但满足一定要求的：以最近RE进行注释
                    else:
                        dis = min(abs(int(pos) - int(start)), abs(int(pos) - int(end)))
                        if dis < min_dis:
                            min_dis = dis
                            nearest_re = origin_re
                        # cnt_match += 1
                        # re_set.add(origin_re)
                        # w.write(origin_re)
                        # w.write('\n')

                # 染色体序号不等，continue，考虑下一个re
                else:
                    cnt_dismatch += 1
                    # 染色体序号不匹配，continue
                    continue
            if flag_fullmatch != 1:
                # 序号匹配，为其注释一个最近RE
                if min_dis < 1000:
                    cnt_unfullmatch += 1
                    print('current min_dis: ', min_dis)
                    re_set.add(nearest_re)
                    # print(min_dis)

    print('cnt_fullmatch : ', cnt_fullmatch)
    print('cnt_unfullmatch : ', cnt_unfullmatch)
    f.close()
    for re in re_set:
        w.write(re)
        w.write('\n')

    w.close()

# 比对数据库提高的糖尿病G，和由RE经过GRN网络得到的G
def T2D_GRN():
    geneORGANizer = datas.import_database_geneORGANizer()
    ovl = 0
    cnt = 0
    set1 = set()
    set2 = set()
    set3 = set()

    for gene in geneORGANizer['genes']:
        set3.add(gene)

    f1 = open('./LiCO/diabetes_mellitus/T2D/SNP2RE2G_5e8.txt', 'r') # 92 884
    f2 = open('./LiCO/diabetes_mellitus/T2D/gene_5e8.txt', 'r') # 459
    w = open('./LiCO/diabetes_mellitus/T2D/884&459_overlap_G.txt', 'w')
    for line1 in f1:
        line1 = line1.split()
        gene1 = line1[0]
        set1.add(gene1)
    print(len(set1))

    for line2 in f2:
        line2 = line2.split()
        gene2 = line2[0]
        set2.add(gene2)
    print(len(set2))

    s = set1.intersection(set2)
    print(len(s))

    for i in s:
        w.write(str(i))
        w.write('\n')

    ss = set2.intersection(set3)
    sss = ss.intersection(set1)
    print(len(ss))
    print(len(sss))

    f1.close()
    f2.close()
    w.close()

# 计算Enrichment ratio
def calc_enrichment_ratio(gene_set, background_genes, enrichment_results):
    # 计算基因集在样本中的显著富集的基因数
    sample_enriched_genes = [gene for gene in gene_set if enrichment_results.get(gene, {}).get('pass_fdr')]
    sample_enriched_genes_count = len(sample_enriched_genes)

    # 计算背景基因组中的显著富集的基因数
    bg_enriched_genes = [gene for gene in background_genes if enrichment_results.get(gene, {}).get('pass_fdr')]
    bg_enriched_genes_count = len(bg_enriched_genes)

    if bg_enriched_genes_count or sample_enriched_genes_count == 0:
        return 0

    # 计算Enrichment ratio
    enrichment_ratio = (float(sample_enriched_genes_count)/len(gene_set)) / (float(bg_enriched_genes_count)/len(background_genes))

    return enrichment_ratio



def pos_sel(): 

    # # 读取Excel文件
    # wb = openpyxl.load_workbook(filename='/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/Top200genes.xlsx')
    # # 获取所有工作表
    # sheets = wb.sheetnames

    # # 读取每个工作表
    # for sheet_name in sheets[1:]:  # 从第2个工作表开始读取
    #     sheet = wb[sheet_name]
    #     species_name = sheet.cell(row=1, column=1).value  # 获取物种名称
    #     gene_names = []
    #     for row in range(3, sheet.max_row+1):  # 从第3行开始读取每一个基因名称
    #         gene_name = sheet.cell(row=row, column=1).value
    #         gene_names.append(gene_name)

    #     # 将每个物种及其基因集合，分别写入本地txt文件中
    #     with open(f'/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/{species_name}.txt', 'w') as f:
    #         print("in write")
    #         f.write(f'Species: {species_name}\n')
    #         f.write('Genes: \n')
    #         for gene_name in gene_names:
    #             f.write(gene_name + '\n')

    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    geneORGANizer = datas.import_database_geneORGANizer()
    vecs = importvec('./LiCO/vec/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')

    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))
    
    Genes_under_selection = []
    cnt = 0
    with open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/P.t. elliotibottom.txt', 'r') as f:
        next(f)
        next(f)
        for line in f:
            line = line.split('\n')
            gene = line[0].upper()
            # print(gene)
            if gene in total_geneset and vecs:
            # if gene in geneORGANizer['genes']:
                Genes_under_selection.append(gene)
                cnt += 1
                # print(gene)
    print('cnt_match_gene: ', cnt)


    fmax, bodyparts = getthres()

    organ_p = dict()
    enrichment_results = {}
    w_cdf = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/No_elliotibottom.txt', 'w')
    for organ in geneORGANizer['bodyparts']:
        if organ not in organ_p:
            organ_p[organ] = 0
        # w_s.write(organ + '\t')
        # w_cdf.write(organ + '\t')
        # w_sf.write(organ + '\t')

        print('curren organ : ', organ)
        # 当前organ在hGraph中的注释情况
        cnt_hG = 0
        cnt_nohG = 0
        cnt_slG = 0
        cnt_noslG = 0
        # for gene in total_geneset:
        for gene in geneORGANizer['genes']:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_hG += 1
                else:
                    cnt_nohG += 1
            else:
                continue

        for gene in Genes_under_selection:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_slG += 1
                else:
                    cnt_noslG += 1
            else:
                continue
        print('cnt_hG : ', cnt_hG)
        print('cnt_slG: ', cnt_slG)

        pv_cdf = 1 - stats.hypergeom.cdf(cnt_slG - 1, len(geneORGANizer['genes']), cnt_hG, len(Genes_under_selection))
        # p_value = scipy.stats.hypergeom.sf(cnt_slG-1, cnt_slG+cnt_noslG+cnt_hG+cnt_nohG, cnt_hG+cnt_nohG, cnt_slG+cnt_noslG)
        enrichment_results[organ] = {'A': cnt_slG, 'B': cnt_noslG, 'C': cnt_hG, 'D': cnt_nohG, 'p_value': pv_cdf}
        # w_cdf.write(str(p_value) + '\n')
    
    # 进行 Bonferroni 矫正
    p_values = np.array([result['p_value'] for result in enrichment_results.values()])
    corrected_p_values = p_values * len(p_values)  # 直接将原始 P 值乘以测试数量

    # 进行 FDR 矫正
    corrected_p_values = list(corrected_p_values)
    pass_threshold = list(multitest.multipletests(corrected_p_values, alpha=0.05, method='fdr_bh')[0])
    pass_organ_labels = list(enrichment_results.keys())

    for i, organ in enumerate(enrichment_results):
        enrichment_results[organ]['pass_fdr'] = corrected_p_values[i]

        w_cdf.write('{}\t'.format(organ))
        w_cdf.write('{}\t'.format(enrichment_results[organ]['p_value']))
        w_cdf.write('{}\n'.format(enrichment_results[organ]['pass_fdr']))

    # 打印通过FDR的GO项
    print("\n通过FDR筛选显著的项：")
    for i, organ in enumerate(pass_organ_labels):
        if pass_threshold[i]:
            print(f"{organ:<15}p-value={p_values[i]:<10.2g}corrected_p_value={corrected_p_values[i]:<10.2g}pass_fdr={pass_threshold[i]}")

    enrichment_ratio = calc_enrichment_ratio(Genes_under_selection, geneORGANizer['genes'], enrichment_results)
    print('Enrichment ratio is:', enrichment_ratio)

    # 获取通过FDR筛选显著的组织列表
    enriched_organ_labels = [organ_label for organ_label in enrichment_results if enrichment_results[organ_label]['pass_fdr']]  

    # 构造可视化所需的数据
    enriched_organ_labels = []
    colors = []
    sizes = []
    for organ_label in enrichment_results:
        if enrichment_results[organ_label]['pass_fdr']:
            enriched_organ_labels.append(organ_label)
            colors.append(-np.log10(enrichment_results[organ_label]['pass_fdr']))
            sizes.append(enrichment_results[organ_label]['A'])

    # 只绘制有显著富集的组织
    if not enriched_organ_labels:
        print('No significant enrichment detected.')
    else:
            # 根据p值，从小到大排序
        sorted_indices = sorted(range(len(p_values)), key=lambda k: p_values[k])
        sorted_organ_labels = [pass_organ_labels[i] for i in sorted_indices if pass_threshold[i]]
        sorted_sizes = [enrichment_results[organ_label]['A'] for organ_label in sorted_organ_labels]
        sorted_p_values = [p_values[i] for i in sorted_indices if pass_threshold[i]]
        sorted_corrected_p_values = [corrected_p_values[i] for i in sorted_indices if pass_threshold[i]]

        # 绘制柱状图
        plt.figure(figsize=(6, 15))
        plt.barh(sorted_organ_labels, sorted_sizes, color='gray')
        plt.gca().invert_yaxis() # 将y轴翻转，让组织名称按富集从上到下排列
        plt.xlabel('organ count')
        plt.ylabel('organ')
        plt.title('Organ count for the gene set')

        plt.savefig('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/enrichment_elliotibottom_barchar.png', dpi=300, bbox_inches='tight')

        # 绘制富集气泡图
        plt.figure(figsize=(6, 15))
        plt.xscale('log')
        sorted_significance = [-np.log10(p) for p in sorted_p_values]
        sns.scatterplot(x=sorted_sizes, y=sorted_organ_labels, hue=sorted_significance, size=sorted_corrected_p_values, sizes=(50, 500))
        plt.legend().remove()
        plt.xlabel('organ count')
        plt.ylabel('organ')
        plt.title('Enriched organ of the gene set')

        # 保存图片
        plt.savefig('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/enrichment_elliotibottom_bubble.png', dpi=300, bbox_inches='tight')

def monkey():
    filename = '/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/ginbon_gene/PAML_pos_gene.xlsx'
    wb = openpyxl.load_workbook(filename)

    # 选择第一个工作表
    sheet = wb.active

    # 定义列表
    data_list = []
    cnt = 0

    # 从第三行开始读取每行的第三列，保存到列表中
    for row in sheet.iter_rows(min_row=3, min_col=3, max_col=8):
        data_list.append(row[0].value)
        cnt += 1
    
    # # 测试结果
    # print(data_list)
    # print(cnt)

    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    geneORGANizer = datas.import_database_geneORGANizer()
    vecs = importvec('./LiCO/vec/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')

    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    cnt_vec = 0
    Genes_under_selection = []
    for gene in data_list:
        if gene in geneORGANizer['genes'] and vecs:
            Genes_under_selection.append(gene)
            cnt_vec += 1
        
    print(f'cnt_vec:{cnt_vec}')


    fmax, bodyparts = getthres()

    organ_p = dict()
    enrichment_results = {}
    w_cdf_pos = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/ginbon_gene/ours_ginbon.txt', 'w')
    w_cdf_neg = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/ginbon_gene/gO_ginbon.txt', 'w')
    for organ in geneORGANizer['bodyparts']:
        if organ not in organ_p:
            organ_p[organ] = 0
        # w_s.write(organ + '\t')
        w_cdf_pos.write(organ + '\t')
        w_cdf_neg.write(organ + '\t')
        # w_sf.write(organ + '\t')

        print('curren organ : ', organ)
        # 当前organ在hGraph中的注释情况
        cnt_hG = 0
        cnt_nohG = 0
        cnt_slG = 0
        cnt_noslG = 0
        # for gene in total_geneset:
        for gene in geneORGANizer['genes']:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_hG += 1
                else:
                    cnt_nohG += 1
            else:
                continue

        for gene in Genes_under_selection:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_slG += 1
                else:
                    cnt_noslG += 1
            else:
                continue
        print('cnt_hG : ', cnt_hG)
        print('cnt_nohG: ', cnt_nohG)
        print('cnt_slG : ', cnt_slG)
        print('cnt_noslG: ', cnt_noslG)



        pv_cdf = 1 - stats.hypergeom.cdf(cnt_slG - 1, len(geneORGANizer['genes']), cnt_hG, len(Genes_under_selection))
        # p_value = scipy.stats.hypergeom.sf(cnt_slG-1, cnt_slG+cnt_noslG+cnt_hG+cnt_nohG, cnt_hG+cnt_nohG, cnt_slG+cnt_noslG)
        # enrichment_results[organ] = {'A': cnt_slG, 'B': cnt_noslG, 'C': cnt_hG, 'D': cnt_nohG, 'p_value': pv_cdf}
        w_cdf_pos.write(str(pv_cdf) + '\n')


        neg_cnt_hG = 0
        neg_cnt_nohG = 0
        neg_cnt_slG = 0
        neg_cnt_noslG = 0
        for gene in geneORGANizer['genes']:
            if gene in vecs:
                if gene in geneORGANizer['bodyparts'][organ]:
                    neg_cnt_hG += 1
                else:
                    neg_cnt_nohG += 1
            else:
                continue

        for gene in Genes_under_selection:
            if gene in vecs:
                if gene in geneORGANizer['bodyparts'][organ]:
                    neg_cnt_slG += 1
                else:
                    neg_cnt_noslG += 1
            else:
                continue
        print('neg_cnt_hG : ', neg_cnt_hG)
        print('neg_cnt_nohG: ', neg_cnt_nohG)
        print('neg_cnt_slG : ', neg_cnt_slG)
        print('neg_cnt_noslG: ', neg_cnt_noslG)

        pv_cdf = 1 - stats.hypergeom.cdf(neg_cnt_slG - 1, len(geneORGANizer['genes']), neg_cnt_hG, len(Genes_under_selection))
        w_cdf_neg.write(str(pv_cdf) + '\n')

    
    # # 进行 Bonferroni 矫正
    # p_values = np.array([result['p_value'] for result in enrichment_results.values()])
    # corrected_p_values = p_values * len(p_values)  # 直接将原始 P 值乘以测试数量

    # # 进行 FDR 矫正
    # corrected_p_values = list(corrected_p_values)
    # pass_threshold = list(multitest.multipletests(corrected_p_values, alpha=0.05, method='fdr_bh')[0])
    # pass_organ_labels = list(enrichment_results.keys())

    # for i, organ in enumerate(enrichment_results):
    #     enrichment_results[organ]['pass_fdr'] = corrected_p_values[i]

    #     w_cdf.write('{}\t'.format(organ))
    #     w_cdf.write('{}\t'.format(enrichment_results[organ]['p_value']))
    #     w_cdf.write('{}\n'.format(enrichment_results[organ]['pass_fdr']))

    # # 打印通过FDR的GO项
    # print("\n通过FDR筛选显著的项：")
    # for i, organ in enumerate(pass_organ_labels):
    #     if pass_threshold[i]:
    #         print(f"{organ:<15}p-value={p_values[i]:<10.2g}corrected_p_value={corrected_p_values[i]:<10.2g}pass_fdr={pass_threshold[i]}")

    # enrichment_ratio = calc_enrichment_ratio(Genes_under_selection, geneORGANizer['genes'], enrichment_results)
    # print('Enrichment ratio is:', enrichment_ratio)

    # # 获取通过FDR筛选显著的组织列表
    # enriched_organ_labels = [organ_label for organ_label in enrichment_results if enrichment_results[organ_label]['pass_fdr']]  

    # # 构造可视化所需的数据
    # enriched_organ_labels = []
    # colors = []
    # sizes = []
    # for organ_label in enrichment_results:
    #     if enrichment_results[organ_label]['pass_fdr']:
    #         enriched_organ_labels.append(organ_label)
    #         colors.append(-np.log10(enrichment_results[organ_label]['pass_fdr']))
    #         sizes.append(enrichment_results[organ_label]['A'])

    # # 只绘制有显著富集的组织
    # if not enriched_organ_labels:
    #     print('No significant enrichment detected.')
    # else:
    #         # 根据p值，从小到大排序
    #     sorted_indices = sorted(range(len(p_values)), key=lambda k: p_values[k])
    #     sorted_organ_labels = [pass_organ_labels[i] for i in sorted_indices if pass_threshold[i]]
    #     sorted_sizes = [enrichment_results[organ_label]['A'] for organ_label in sorted_organ_labels]
    #     sorted_p_values = [p_values[i] for i in sorted_indices if pass_threshold[i]]
    #     sorted_corrected_p_values = [corrected_p_values[i] for i in sorted_indices if pass_threshold[i]]

    #     # 绘制柱状图
    #     plt.figure(figsize=(6, 15))
    #     plt.barh(sorted_organ_labels, sorted_sizes, color='gray')
    #     plt.gca().invert_yaxis() # 将y轴翻转，让组织名称按富集从上到下排列
    #     plt.xlabel('organ count')
    #     plt.ylabel('organ')
    #     plt.title('Organ count for the gene set')

    #     plt.savefig('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/enrichment_elliotibottom_barchar.png', dpi=300, bbox_inches='tight')

    #     # 绘制富集气泡图
    #     plt.figure(figsize=(6, 15))
    #     plt.xscale('log')
    #     sorted_significance = [-np.log10(p) for p in sorted_p_values]
    #     sns.scatterplot(x=sorted_sizes, y=sorted_organ_labels, hue=sorted_significance, size=sorted_corrected_p_values, sizes=(50, 500))
    #     plt.legend().remove()
    #     plt.xlabel('organ count')
    #     plt.ylabel('organ')
    #     plt.title('Enriched organ of the gene set')

    #     # 保存图片
    #     plt.savefig('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/pos_sel/enrichment_elliotibottom_bubble.png', dpi=300, bbox_inches='tight')

def imprinted():
    filename = '/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/imprinted_gene/Imprinted genes.txt'
    # 定义列表
    data_list = []
    cnt = 0
    with open(filename) as f:
        next(f)
        for line in f:
            line = line.split()
            gene = line[0]
            data_list.append(gene)
            cnt += 1

    
    # 测试结果
    # print(data_list)
    print(cnt)

    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    geneORGANizer = datas.import_database_geneORGANizer()
    vecs = importvec('./LiCO/vec/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')

    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    cnt_vec = 0
    Genes_under_selection = []
    for gene in data_list:
        if gene in geneORGANizer['genes'] and vecs:
            Genes_under_selection.append(gene)
            cnt_vec += 1
        
    print(f'cnt_vec:{cnt_vec}')


    fmax, bodyparts = getthres()

    organ_p = dict()
    enrichment_results = {}
    w_cdf_pos = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/imprinted_gene/ours_imprinted_gene.txt', 'w')
    w_cdf_neg = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/imprinted_gene/gO_imprinted_gene.txt', 'w')
    for organ in geneORGANizer['bodyparts']:
        if organ not in organ_p:
            organ_p[organ] = 0
        # w_s.write(organ + '\t')
        w_cdf_pos.write(organ + '\t')
        w_cdf_neg.write(organ + '\t')
        # w_sf.write(organ + '\t')

        print('curren organ : ', organ)
        # 当前organ在hGraph中的注释情况
        cnt_hG = 0
        cnt_nohG = 0
        cnt_slG = 0
        cnt_noslG = 0
        # for gene in total_geneset:
        for gene in geneORGANizer['genes']:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_hG += 1
                else:
                    cnt_nohG += 1
            else:
                continue

        for gene in Genes_under_selection:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_slG += 1
                else:
                    cnt_noslG += 1
            else:
                continue
        print('cnt_hG : ', cnt_hG)
        print('cnt_nohG: ', cnt_nohG)
        print('cnt_slG : ', cnt_slG)
        print('cnt_noslG: ', cnt_noslG)



        pv_cdf = 1 - stats.hypergeom.cdf(cnt_slG - 1, len(geneORGANizer['genes']), cnt_hG, len(Genes_under_selection))
        # p_value = scipy.stats.hypergeom.sf(cnt_slG-1, cnt_slG+cnt_noslG+cnt_hG+cnt_nohG, cnt_hG+cnt_nohG, cnt_slG+cnt_noslG)
        # enrichment_results[organ] = {'A': cnt_slG, 'B': cnt_noslG, 'C': cnt_hG, 'D': cnt_nohG, 'p_value': pv_cdf}
        w_cdf_pos.write(str(pv_cdf) + '\n')


        neg_cnt_hG = 0
        neg_cnt_nohG = 0
        neg_cnt_slG = 0
        neg_cnt_noslG = 0
        for gene in geneORGANizer['genes']:
            if gene in vecs:
                if gene in geneORGANizer['bodyparts'][organ]:
                    neg_cnt_hG += 1
                else:
                    neg_cnt_nohG += 1
            else:
                continue

        for gene in Genes_under_selection:
            if gene in vecs:
                if gene in geneORGANizer['bodyparts'][organ]:
                    neg_cnt_slG += 1
                else:
                    neg_cnt_noslG += 1
            else:
                continue
        print('neg_cnt_hG : ', neg_cnt_hG)
        print('neg_cnt_nohG: ', neg_cnt_nohG)
        print('neg_cnt_slG : ', neg_cnt_slG)
        print('neg_cnt_noslG: ', neg_cnt_noslG)

        pv_cdf = 1 - stats.hypergeom.cdf(neg_cnt_slG - 1, len(geneORGANizer['genes']), neg_cnt_hG, len(Genes_under_selection))
        w_cdf_neg.write(str(pv_cdf) + '\n')

def new_echo():
    filename = '/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/echo/gene.txt'
    # 定义列表
    data_list = []
    cnt = 0
    with open(filename) as f:
        next(f)
        for line in f:
            line = line.split()
            gene = line[0]
            data_list.append(gene)
            cnt += 1

    
    # 测试结果
    # print(data_list)
    print(cnt)

    grn = datas.import_database_grnhg()
    ppi = datas.importppi()
    geneORGANizer = datas.import_database_geneORGANizer()
    vecs = importvec('./LiCO/vec/' + '10_15_7path_addGB_4layerTrain_humanGRN_2_vec.txt')

    total_geneset = (set(geneORGANizer['genes'].keys()).union(set(ppi.keys()))).union(set(grn['gene_tfs'].keys()))

    cnt_vec = 0
    Genes_under_selection = []
    for gene in data_list:
        if gene in geneORGANizer['genes'] and vecs:
            Genes_under_selection.append(gene)
            cnt_vec += 1
        
    print(f'cnt_vec:{cnt_vec}')


    fmax, bodyparts = getthres()

    organ_p = dict()
    enrichment_results = {}
    w_cdf_pos = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/echo/ours_echo.txt', 'w')
    w_cdf_neg = open('/share/home/liangzhongming/NetOrganizer/RE-ORGANizer-master/generateREGOA/NO_plusAlc/echo/gO_echo.txt', 'w')
    for organ in geneORGANizer['bodyparts']:
        if organ not in organ_p:
            organ_p[organ] = 0
        # w_s.write(organ + '\t')
        w_cdf_pos.write(organ + '\t')
        w_cdf_neg.write(organ + '\t')
        # w_sf.write(organ + '\t')

        print('curren organ : ', organ)
        # 当前organ在hGraph中的注释情况
        cnt_hG = 0
        cnt_nohG = 0
        cnt_slG = 0
        cnt_noslG = 0
        # for gene in total_geneset:
        for gene in geneORGANizer['genes']:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_hG += 1
                else:
                    cnt_nohG += 1
            else:
                continue

        for gene in Genes_under_selection:
            if gene in vecs:
                if cosine_similarity([vecs[gene], vecs[organ]])[0][1] >= fmax[organ]:
                    cnt_slG += 1
                else:
                    cnt_noslG += 1
            else:
                continue
        print('cnt_hG : ', cnt_hG)
        print('cnt_nohG: ', cnt_nohG)
        print('cnt_slG : ', cnt_slG)
        print('cnt_noslG: ', cnt_noslG)



        pv_cdf = 1 - stats.hypergeom.cdf(cnt_slG - 1, len(geneORGANizer['genes']), cnt_hG, len(Genes_under_selection))
        # p_value = scipy.stats.hypergeom.sf(cnt_slG-1, cnt_slG+cnt_noslG+cnt_hG+cnt_nohG, cnt_hG+cnt_nohG, cnt_slG+cnt_noslG)
        # enrichment_results[organ] = {'A': cnt_slG, 'B': cnt_noslG, 'C': cnt_hG, 'D': cnt_nohG, 'p_value': pv_cdf}
        w_cdf_pos.write(str(pv_cdf) + '\n')


        neg_cnt_hG = 0
        neg_cnt_nohG = 0
        neg_cnt_slG = 0
        neg_cnt_noslG = 0
        for gene in geneORGANizer['genes']:
            if gene in vecs:
                if gene in geneORGANizer['bodyparts'][organ]:
                    neg_cnt_hG += 1
                else:
                    neg_cnt_nohG += 1
            else:
                continue

        for gene in Genes_under_selection:
            if gene in vecs:
                if gene in geneORGANizer['bodyparts'][organ]:
                    neg_cnt_slG += 1
                else:
                    neg_cnt_noslG += 1
            else:
                continue
        print('neg_cnt_hG : ', neg_cnt_hG)
        print('neg_cnt_nohG: ', neg_cnt_nohG)
        print('neg_cnt_slG : ', neg_cnt_slG)
        print('neg_cnt_noslG: ', neg_cnt_noslG)

        pv_cdf = 1 - stats.hypergeom.cdf(neg_cnt_slG - 1, len(geneORGANizer['genes']), neg_cnt_hG, len(Genes_under_selection))
        w_cdf_neg.write(str(pv_cdf) + '\n')

if __name__ == '__main__':
    # altitude()
    # generateDict_gene_predict_bodypart()
    # gO_B_link_G()
    # link_b()
    # findSpecificTFs()
    echo()
    enrichment()
    FDR()
    # diabetes()
    # enrichment_re()
    # FDR()
    # ad()
    # enrichment_re()
    # FDR()
    # import_T2D_diabetes_gene()
    # enrichment()
    # FDR()

    # import_GWAS()
    # enrichment_re()

    # import_GWAS()
    # import_GWAS_genes()
    # enrichment()


    # import_T2D_diabetes_gene()
    # diabetes()
    # T2D_GRN()
    # import_T2D_diabetes_gene()


    # calSimAvg()

    # pos_sel()
    monkey()
    imprinted()
    # new_echo()

    # enrichment()
from fnmatch import fnmatch
import pandas as pd


gO_organs = ['cerebellum','brain','head','uterus','ear','outer ear','middle ear','inner ear','peripheral nerve','peripheral nervous system','cranial nerve','pituitary gland','ovary','testicle','penis','vagina','vulva','skin','hair','prostate','breast','eye','lower limb','spinal cord','mouth','skull','spinal column','vertebrae','pelvis','rib','rib cage','lung','foot','tibia','fibula','femur','sternum','chest wall','clavicle','scapula','humerus','radius','ulna','neck','mandible','maxilla','jaw','ankle','knee','hip','wrist','elbow','shoulder','finger','toe','digit','hand','forearm','arm','shin','thigh','upper limb','face','eyelid','kidney','urethra','heart','anus','urinary bladder','blood vessel','pancreas','bronchus','white blood cell','blood','liver','biliary tract','gallbladder','olfactory bulb','ureter','red blood cell','coagulation system','aorta','heart valve','rectum','tooth','nose','intestine','large intestine','forehead','diaphragm','stomach','small intestine','chin','cheek','bone marrow','lip','pharynx','nail','meninges','cerebrospinal fluid','spleen','lymph node','sinus','abdominal wall','duodenum','esophagus','placenta','larynx','trachea','vocal cord','epiglottis','tongue','scalp','lymph vessel','lacrimal apparatus','adrenal gland','sweat gland','salivary gland','thyroid','hypothalamus','appendix','parathyroid','thymus','fallopian tube','vas deferens']
gO_regions = ['limbs', 'head and neck', 'thorax', 'abdomen', 'pelvis', 'general']
gO_systems = ['immune', 'cardiovascular', 'nervous', 'skeleton', 'skeletal muscle', 'reproductive', 'digestive',	'urinary', 'respiratory', 'endocrine', 'lymphatic', 'integumentary']
gO_germlayers = ['endoderm', 'mesoderm', 'ectoderm']
# def importgoa(go,filename):
#     genes=dict()
#     terms=dict()
#     with open (filename,'r') as f:
#         # 根据filename中的内容，生成genes和terms两个词典，分别是(gene,term)和(term,gene)对应
#         for line in f:
#             line = line.split()
#             gene = line[0]
#             term = line[1]
#             if term in go:
#                 if not gene in genes:
#                    genes[gene]=set()
#                 genes[gene].add(term)
#                 if not term in terms:
#                     terms[term]=set()
#                 terms[term].add(gene)
#
#     return {'genes':genes,'terms':terms}#,'genetpr':genetpr,'termtpr':termtpr

# def goatpr(goa,go): #true path rule
#     genes=goa['genes']
#     terms=goa['terms']
#     genetpr=dict()
#     termtpr=dict()
#     # 依据terms和genes中的内容，生成词典termtpr和genetpr
#     # term只对应一个gene
#     for term in terms.keys():
#         termtpr[term] = terms[term]
#     # gene会对用多个term
#     for gene in genes.keys():
#         # 将genes中的每个gene作为genetpr中的key
#         genetpr[gene] = genes[gene]
#         # 对genes中每个gene对应的term
#         for term in list(genes[gene]):
#             # 若该term在go[term][ancesent]中，则将该term纳入genetpr中，作为当前gene对应的value ?
#             for t in go[term]['ancesent']:
#                 genetpr[gene].add(t)
#                 # 若同时该term未在termtpr中，则创建，并纳入该term对应的gene
#                 if not t in termtpr:
#                     termtpr[t]=set()
#                 termtpr[t].add(gene)
#     return {'genes':genetpr,'terms':termtpr}#  'genetpr':genetpr,'termtpr':termtpr
import openpyxl
from openpyxl import load_workbook

def importgrnhg(filename='./datahg/grn-human.txt'):
    genes = dict()
    res = dict()
    tfs = dict()
    with open(filename, 'r') as f:
        next(f)
        for line in f:
            line = line.split()
            gene = line[0]
            re = line[1]
            tf = line[2]
            if not gene in genes:
                genes[gene] = set()
            genes[gene].add(re)
            if not re in res:
                res[re] = {'reg': set(), 'binded': set()}
            res[re]['reg'].add(gene)
            res[re]['binded'].add(tf)
            if not tf in tfs:
                tfs[tf] = set()
            tfs[tf].add(re)
    return {'genes': genes, 'res': res, 'tfs': tfs}

def importgrnmgi(filename='./datamgi/grn-mouse.txt'):
    genes = dict()
    res = dict()
    tfs = dict()
    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            re = line[2]
            tf = line[-2].split(';')
            if not gene in genes:
                genes[gene] = set()
            genes[gene].add(re)
            if not re in res:
                res[re] = {'reg': set(), 'binded': set()}
            res[re]['reg'].add(gene)
            for tff in tf:
                res[re]['binded'].add(tff)
                if not tff in tfs:
                    tfs[tff] = set()
                tfs[tff].add(re)
    return {'genes': genes, 'res': res, 'tfs': tfs}

# def importgo(filename):
#     # Reading Gene Ontology from OBO Formatted file
#     go = dict()
#     obj = None
#     ns={'biological_process':'BP','molecular_function':'MF','cellular_component':'CC'}
#     with open(filename, 'r') as f:
#         for line in f:
#             line = line.strip()  # 移除字符串头尾指定字符，默认为空格/换行符
#             if not line:
#                 continue
#             if line == '[Term]':
#                 if obj is not None:
#                     go[obj['id']] = obj
#                 obj = dict()
#                 obj['is_a'] = set()
#                 obj['part_of'] = set()
#                 obj['regulates'] = set()
#                 obj['is_obsolete'] = False
#                 obj['ancesent'] = set()
#                 obj['descent'] = set()
#                 continue
#             elif line == '[Typedef]':
#                 if obj is not None:
#                     go[obj['id']] = obj
#                 obj = None
#             else:
#                 if obj is None:
#                     continue
#                 l = line.split(": ")
#                 if l[0] == 'id':
#                     obj['id'] = l[1]
#                 elif l[0] == 'is_a':
#                     obj['is_a'].add(l[1].split(' ! ')[0])
#                 elif l[0] == 'name':
#                     obj['name'] = l[1]
#                 elif l[0] == 'is_obsolete' and l[1] == 'true':
#                     obj['is_obsolete'] = True
#                 elif l[0] == 'namespace':
#                     obj['namespace']=ns[l[1]]
#     if obj is not None:
#         go[obj['id']] = obj
#     for go_id in list(go.keys()):
#         if go[go_id]['is_obsolete']:
#             del go[go_id]
#     for go_id, val in go.items():
#         if 'children' not in val:
#             val['children'] = set()
#         for p_id in val['is_a']:
#             if p_id in go:
#                 if 'children' not in go[p_id]:
#                     go[p_id]['children'] = set()
#                 go[p_id]['children'].add(go_id)
#     for go_id in go.keys():
#         if 'ancesent' not in go[go_id]:
#             go[go_id]['ancesent']=set()
#         temp = list(go[go_id]['is_a'])
#         while temp:
#             now=temp.pop()
#             go[go_id]['ancesent'].add(now)
#             temp = list(set(temp)|go[now]['is_a'])
#             if 'descent' not in go[now]:
#                 go[now]['descent']=set()
#             go[now]['descent'].add(go_id)
#     return go

def importgene_organs(filename='./New_gene_organs_2.txt'):
    genes = dict()
    organs = dict()
    condi_genes = set()
    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0].upper() # 转换为大写
            # print(gene)
            if not gene in genes:
                genes[gene] = set()
            for index in range(1, 126):  # 下标1~125存放organs是否被影响的flag
                organ = gO_organs[index - 1]
                if not organ in organs:
                    organs[organ] = set()
                if line[index] == '1':  # line[1]~line[125]

                    genes[gene].add(organ)
                    organs[organ].add(gene)
                    # print(organs)


    # f.close()
    # print(genes)
    # print(organs['vas deferens'])
    # print(len(organs.keys()))
    # print(organs[])

    # 顺带由New_gene_organs_2.txt邻接矩阵的形式，生成neg.txt和pos.txt
    fneg = open('./datamgi/pos_neg/neg.txt', 'w')
    count = 0
    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            # print(len(line))
            for index in range(1, 126):
                # print('index: ' + str(index))
                # print('line[index]: ' + str(line[index]))
                if line[index] == '0':
                    count += 1
                    # print(gene, ' count: ', count)
            for index in range(1, 126):
                if count == 125:  # 该基因与任何组织都无关,不考虑
                    with open('./datamgi/pos_neg/abandon_gene.txt', 'a+') as fa:
                        fa.write(str(gene) + '\n')
                        count = 0
                    break
                else:
                    if line[index] == '0':  # 构造负样本
                        organ = gO_organs[index - 1]
                        fneg.write(gene + '\t')
                        fneg.write(organ + '\n')
                        condi_genes.add(gene)
                    count = 0

    fpos = open('./datamgi/pos_neg/pos.txt', 'w')
    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            # print(len(line))
            for index in range(1, 126):
                # print('index: ' + str(index))
                # print('line[index]: ' + str(line[index]))
                if line[index] == '1':  # 构造正样本
                    organ = gO_organs[index - 1]
                    fpos.write(gene + '\t')
                    fpos.write(organ + '\n')


    return {'genes': genes, 'organs': organs, 'condi_genes':condi_genes}


def importgeneORGANizer():
    genesOfOrgan = dict()
    genesOfRegion = dict()
    genesOfSystem = dict()
    genesOfGermlayer = dict()

    organs = dict()
    regions = dict()
    systems = dict()
    germlayers = dict()

    # organ
    with open('./gene_organ.txt', 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            if not gene in genesOfOrgan:
                genesOfOrgan[gene] = set()
            for index in range(1, 126):  # 下标1~125存放organs是否被影响的flag
                organ = gO_organs[index - 1]
                if not organ in organs:
                    organs[organ] = set()
                if line[index] == '1':  # line[1]~line[125]

                    genesOfOrgan[gene].add(organ)
                    organs[organ].add(gene)
    # region
    with open('./gene_region.txt', 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            if not gene in genesOfRegion:
                genesOfRegion[gene] = set()
            for index in range(1, 7):  # 下标1~XXX存放regions是否被影响的flag
                region = gO_regions[index - 1]
                if not region in regions:
                    regions[region] = set()
                if line[index] == '1':  # line[1]~line[125]

                    genesOfRegion[gene].add(region)
                    regions[region].add(gene)

    # system
    with open('./gene_system.txt', 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            if not gene in genesOfSystem:
                genesOfSystem[gene] = set()
            for index in range(1, 13):  # 下标1~XXX存放systems是否被影响的flag
                system = gO_systems[index - 1]
                if not system in systems:
                    systems[system] = set()
                if line[index] == '1':  # line[1]~line[125]

                    genesOfSystem[gene].add(system)
                    systems[system].add(gene)

    # germlayer
    with open('./gene_germlayer.txt', 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            if not gene in genesOfGermlayer:
                genesOfGermlayer[gene] = set()
            for index in range(1, 4):  # 下标1~XXX存放germlayers是否被影响的flag
                germlayer = gO_germlayers[index - 1]
                if not germlayer in germlayers:
                    germlayers[germlayer] = set()
                if line[index] == '1':  # line[1]~line[125]

                    genesOfGermlayer[gene].add(germlayer)
                    germlayers[germlayer].add(gene)

    return {'genesOfOrgan': genesOfOrgan, 'genesOfRegion': genesOfRegion, 'genesOfSystem': genesOfSystem, 'genesOfGermlayer': genesOfGermlayer, 'organs': organs, 'systems': systems, 'regions': regions, 'germlayers': germlayers}

def importgeneItem(filename='./model0.87/data/DX_train_total4Layer.txt'):
    genes = dict()
    bodyparts = dict()
    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            if not gene in genes:
                genes[gene] = set()
            bodypart = line[1]
            for index in range(2, len(line)):
                if line[index].isalpha():
                    bodypart += ' '
                    bodypart += line[index]
                else:
                    break
            if not bodypart in bodyparts:
                bodyparts[bodypart] = set()
            genes[gene].add(bodypart)
            bodyparts[bodypart].add(gene)

    return {'genes': genes, 'bodyparts': bodyparts}

# def importgene_germlayers(filename):
#     g_g = dict()
#     with open(filename, 'r') as f:
#         for line in f:
#             line = line.split()
#             if not line[0] in g_g:
#                 g_g[line[0]] = dict()
#             if not line[1] in g_g:
#                 g_g[line[1]] = dict()
#     return g_g
#
#
# def importgene_regions(filename):
#     g_r = dict()
#     with open(filename, 'r') as f:
#         for line in f:
#             line = line.split()
#             if not line[0] in g_r:
#                 g_r[line[0]] = dict()
#             if not line[1] in g_r:
#                 g_r[line[1]] = dict()
#     return g_r
#
# def importgene_systems(filename):
#     g_s = dict()
#     with open(filename, 'r') as f:
#         for line in f:
#             line = line.split()
#             if not line[0] in g_s:
#                 g_s[line[0]] = dict()
#             if not line[1] in g_s:
#                 g_s[line[1]] = dict()
#     return g_s

def importppi(filename='./datahg/ppifull.txt'):
    s=dict()
    # 导入蛋白互作网络PPI 包括互作的两个蛋白和他们的权重
    with open(filename,'r') as f:
        for line in f:
            line=line.split()
            if not line[0] in s:
                s[line[0]]=dict()
            if not line[1] in s:
                s[line[1]]=dict()
            weight = int(line[2])
            weight = 1
            s[line[0]][line[1]]=weight
            s[line[1]][line[0]]=weight
    return s

def importbodyPartLinks(filename='./process_result/bodypart_links.txt'):
    bodyparts = dict()

    with open(filename, 'r') as f:
        for line in f:
            # 删除换行符
            line = line.rstrip('\n')
            line = line.split('/')
            bodypart = line[0]
            # print(line[0])
            if not bodypart in bodyparts:
                bodyparts[bodypart] = set()
            i = 1
            while True:
                if line[i:] == []:
                    break
                otherbp = line[i]
                i += 1
                bodyparts[bodypart].add(otherbp)
    # print(bodyparts)
    # print(bodyparts['eye'])

    return bodyparts



def import_database_TRRUST(filename='./model0.87/data/trrust_rawdata.human.tsv'):
    tf_dict = dict()
    tg_dict = dict()
    count_tf = 0
    count_tg = 0
    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            tf = line[0]
            tg = line[1]
            if tf not in tf_dict:
                count_tf += 1
                tf_dict[tf] = set()
            tf_dict[tf].add(tg)

            if tg not in tg_dict:
                count_tg += 1
                tg_dict[tg] = set()
            tg_dict[tg].add(tf)
    # print('tg : ', count_tg)
    # print('tf : ', count_tf)
    return {'tfs': tf_dict, 'tgs': tg_dict}


def import_database_geneORGANizer_organ(filename='./model0.87/data/new_gene_organName.txt'):
    genes = dict()
    organs = dict()

    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0].upper()
            if gene not in genes:
                genes[gene] = set()
            organ = line[1]
            for index in range(2, len(line)):
                if line[index].isalpha():
                    organ += ' '
                    organ += line[index]
                else:
                    break
            if organ not in organs:
                organs[organ] = set()
            genes[gene].add(organ)
            organs[organ].add(gene)

    return {'genes': genes, 'organs': organs}

def import_database_geneORGANizer(filename='/share/home/liangzhongming/RE_ORGANize_4layers/RE-ORGANizer-master/generateREGOA/model0.87/data/DX_pos.txt'):
    genes = dict()
    bodyparts = dict()

    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            gene = line[0]
            if not gene in genes:
                genes[gene] = set()
            bodypart = line[1]
            for index in range(2, len(line)):
                if line[index].isalpha():
                    bodypart += ' '
                    bodypart += line[index]
                else:
                    break
            if not bodypart in bodyparts:
                bodyparts[bodypart] = set()
            genes[gene].add(bodypart)
            bodyparts[bodypart].add(gene)

    return {'genes': genes, 'bodyparts': bodyparts}

def import_database_grnhg(filename='./datahg/grn-human.txt'):
    gene_tf = dict()
    gene_re = dict()
    res = dict()
    tfs = dict()
    tf_res = dict()
    count_tf = 0
    count_tg = 0
    count_re = 0
    with open(filename, 'r') as f:
        next(f)
        for line in f:
            line = line.split()
            gene = line[0]
            re = line[1]
            tf = line[2]

            if gene not in gene_tf:
                count_tg += 1
                gene_tf[gene] = set()
            gene_tf[gene].add(tf)

            if gene not in gene_re:
                count_re += 1
                gene_re[gene] = set()
            gene_re[gene].add(re)

            if not re in res:
                res[re] = {'reg': set(), 'binded': set()}
            res[re]['reg'].add(gene)
            res[re]['binded'].add(tf)

            if tf not in tfs:
                count_tf += 1
                tfs[tf] = set()
            tfs[tf].add(gene)

            if tf not in tf_res:
                tf_res[tf] = set()
            tf_res[tf].add(re)
    # print('tg in GRN, not sure in vecs: ', count_tg)
    # print('tf : ', count_tf)
    return {'tfs': tfs, 'gene_tfs': gene_tf, 'res': res, 'gene_res': gene_re, 'tf_res': tf_res}

def import_database_cellMarker(filename = './LiCO/data/human_cell_marker.txt'):
    normal_organs_dict = dict()
    cancer_organs_dict = dict()
    normal_genes_dict = dict()
    cancer_genes_dict = dict()

    wb = openpyxl.load_workbook('./LiCO/data/human_cellMarker.xlsx')
    ws = wb.active

    with open('./LiCO/data/normale_cellmarker.txt', 'w') as fwn, \
         open('./LiCO/data/cancer_cellmarker.txt', 'w') as fwc:

        maxrows = ws.max_row
        for line in ws.iter_rows(min_row=2): # 忽略第一行
            count = 0
            for cell in line:
                count += 1
                if count == 2:
                    organ = cell.value.lower()
                    # print('organ : ' + organ)
                if count == 5:
                    state = cell.value
                    # print('state : ' + state)
                if count == 8:
                    marker = cell.value
                    # print('marker : ' + str(marker))

                    if state == 'Normal cell':
                        if organ not in normal_organs_dict:
                            normal_organs_dict[organ] = set()
                        # fwn.write(organ + '\t')

                        # 对marker进行拆分
                        marker = str(marker)
                        marker = marker.split(', ')
                        for idx in range(len(marker)):
                            gene = marker[idx]
                            if gene not in normal_genes_dict:
                                normal_genes_dict[gene] = set()
                            normal_genes_dict[gene].add(organ)
                            normal_organs_dict[organ].add(gene)
                        #     fwn.write(str(marker[idx]) + '\t')
                        # fwn.write('\n')
                    elif state == 'Cancer cell':
                        if organ not in cancer_organs_dict:
                            cancer_organs_dict[organ] = set()
                        # fwc.write(organ + '\t')

                        marker = str(marker)
                        marker = marker.split(', ')
                        for idx in range(len(marker)):
                            gene = marker[idx]
                            if gene not in cancer_genes_dict:
                                cancer_genes_dict[gene] = set()
                            cancer_genes_dict[gene].add(organ)
                            cancer_organs_dict[organ].add(marker[idx])
                        #     fwc.write(str(marker[idx]) + '\t')
                        # fwc.write('\n')

    return {'normal_organs': normal_organs_dict, 'cancer_organs': cancer_organs_dict, 'cancer_genes': cancer_genes_dict, 'normal_genes': normal_genes_dict}




    # with open(filename, 'r') as f, \
    #      open('./LiCO/data/normale_cellmarker.txt', 'w') as fwn, \
    #      open('./LiCO/data/cancer_cellmarker.txt', 'w') as fwc:
    #     next(f)
    #     for line in f:
    #         # speciesType	tissueType	UberonOntologyID	cancerType	cellType	cellName	CellOntologyID	cellMarker	geneSymbol	geneID	proteinName	proteinID	markerResource	PMID	Company
    #         line = line.split()
    #         organ = line[1].lower()
    #         print('organ : ' + organ)
    #         marker_genes = line[7]
    #         print('marker_genes : ' + marker_genes)
    #         print('\n')
    #         if line[4] == 'Normal cell':
    #             if organ not in normal_organs_dict:
    #                 normal_organs_dict[organ] = set()
    #                 fwn.write(organ + '\t')
    #                 # add cellMarker
    #                 # 形式：以','分隔
    #             marker_genes = marker_genes.split(',')
    #             for idx in range(0, len(marker_genes)):
    #                 normal_organs_dict[organ].add(marker_genes[idx])
    #                 fwn.write(marker_genes[idx] + '\t')
    #             fwn.write('\n')
    #
    #         else:
    #             if organ not in cancer_organs_dict:
    #                 cancer_organs_dict[organ] = set()
    #                 fwc.write(organ + '\t')
    #
    #             marker_genes = marker_genes.split(',')
    #             for idx in range(0, len(marker_genes)):
    #                 normal_organs_dict[organ].add(marker_genes[idx])
    #                 fwc.write(marker_genes[idx] + '\t')
    #             fwc.write('\n')

def import_database_RNA_Atlas(filename = './LiCO/data/RNA_Seq_Atlas_rev1.txt'):
    RNA_Atlas_organs = ['adipose', 'large intestine', 'heart', 'hypothalamus', 'kidney', 'liver', 'lung', 'ovary', 'skeletal muscle', 'spleen', 'testicle']
    with open(filename, 'r') as f:
        genes = dict()
        organs = dict()

        count_genes = 0
        next(f)
        # entrez_gene_id	ensembl_gene_id	hgnc_symbol	transcript	transcript_length	adipose	colon	heart	hypothalamus	kidney	liver	lung	ovary	skeletalmuscle	spleen	testes
        for line in f:
            line = line.split()
            if fnmatch(line[2], '*_*'): # 使用通配符匹配
                gene = line[1]
                if gene not in genes:
                    genes[gene] = set()
                    count_genes += 1
                for index in range(4, len(line)):

                    if not fnmatch(line[index], '0'):
                        organ = RNA_Atlas_organs[index - 5]
                        if organ not in organs:
                            organs[organ] = set()
                        organs[organ].add(gene)

                        genes[gene].add(organ)
            else:
                gene = line[2]
                if gene not in genes:
                    genes[gene] = set()
                    count_genes += 1
                for index in range(5, len(line)):

                    if not fnmatch(line[index], '0'):
                        organ = RNA_Atlas_organs[index - 5]
                        if organ not in organs:
                            organs[organ] = set()
                        organs[organ].add(gene)
                        genes[gene].add(organ)
    # print('RNA-Atlas gene : ', count_genes)
    return {'genes': genes, 'organs': organs}


def import_database_hTFtarget(filename = './LiCO/data/TF-Target-information.txt'):
    with open(filename, 'r') as f:
        gene_tfs = dict()
        tf_genes = dict()
        tissue_genes = dict()
        tissue_tfs = dict()


        count_genes = 0
        count_tfs = 0
        count_tissues = 0
        next(f)
       # TF	target	tissue
        for line in f:
            line = line.split()
            tf = line[0]
            gene = line[1]

            if tf not in tf_genes:
                count_tfs += 1
                tf_genes[tf] = set()
            tf_genes[tf].add(gene)

            if gene not in gene_tfs:
                count_genes += 1
                gene_tfs[gene] = set()
            gene_tfs[gene].add(tf)

            tissue_set = line[2]
            tissue_set = tissue_set.split(',')
            for index in range(len(tissue_set)):
                tissue = tissue_set[index]

                if tissue not in tissue_genes:
                    # print(tissue)
                    count_tissues += 1
                    tissue_genes[tissue] = set()
                tissue_genes[tissue].add(gene)

                if tissue not in tissue_tfs:
                    tissue_tfs[tissue] = set()
                tissue_tfs[tissue].add(tf)


    print('hTFtarget gene : ', count_genes)
    print('hTFtarget tf : ', count_tfs)
    print('hTFtarget tissue : ', count_tissues)

    return {'genes': gene_tfs, 'tfs': tf_genes, 'tissue_tfs': tissue_tfs, 'tissue_genes': tissue_genes}

def import_predict_TF_TG_fmax(filename = './LiCO/threshold/grn_tf_tg_fmax.txt'):
    tf_dict = dict()
    tg_dict = dict()

    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            tf = line[0]
            if tf not in tf_dict:
                tf_dict[tf] = set()

            for idx in range(1, len(line)):
                tg = line[idx]
                if tg not in tg_dict:
                    tg_dict[tg] = set()
                tf_dict[tf].add(tg)
                tg_dict[tg].add(tf)

    return {'tfs': tf_dict, 'tgs': tg_dict}


def import_predict_b_g_fmax(filename = './LiCO/threshold/gO_bodypart_G_fmax.txt'):
    b_dict = dict()
    g_dict = dict()

    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            b = line[0]
            # 格式：tissueName   geneName 前者为小写 后者为大写
            for index in range(1, len(line)):  # organs名字可能是'xxx xxx xxx'形式
                if line[index].islower():
                    b += ' '
                    b += line[index]
                else:
                    break

            if b not in b_dict:
                b_dict[b] = set()

            for idx in range(index, len(line)):
                g = line[idx]
                if g not in g_dict:
                    g_dict[g] = set()
                b_dict[b].add(g)
                g_dict[g].add(b)

    return {'bodyparts': b_dict, 'genes': g_dict}

def import_predict_b_re_fmax(filename = './LiCO/threshold/gO_bodypart_RE_fmax.txt'):
    b_dict = dict()
    re_dict = dict()

    with open(filename, 'r') as f:
        for line in f:
            line = line.split()
            b = line[0]
            # 格式：tissueName   RE
            for index in range(1, len(line)):  # organs名字可能是'xxx xxx xxx'形式
                if line[index].isalpha():
                    b += ' '
                    b += line[index]
                else:
                    break

            if b not in b_dict:
                b_dict[b] = set()
            for idx in range(index, len(line)):
                re = line[idx]
                if re.isalpha():
                    break

                if re not in re_dict:
                    re_dict[re] = set()
                b_dict[b].add(re)
                re_dict[re].add(b)
    # print(len(b_dict.keys()))
    # print(b_dict.keys())
    # print(len(re_dict.keys()))

    return {'bodyparts': b_dict, 'res': re_dict}

def read_GTEx_2_dict(filename='./LiCO/data/GTEx_tissue_spe_eQTL/'):
    count = 0
    with open('./LiCO/data/GTEx.txt', 'w') as w:
        w.write('gene' + '\t' + 'pos' + '\t' + 'tissue') # 将tissue放最后 方便拿 否则tissue name 和 gene name不好区分开
        w.write('\n')

        # 读取37个tissu的文件,去重后共23个tissue
        filename_list = [filename+'Adrenal_Gland_Analysis_cis-eQTLs.txt',
                         filename + 'Artery_Aorta_Analysis_cis-eQTLs.txt',
                         filename + 'Artery_Coronary_Analysis_cis-eQTLs.txt',
                         filename + 'Artery_Tibial_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Anterior_cingulate_cortex_BA24_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Caudate_basal_ganglia_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Cerebellar_Hemisphere_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Cerebellum_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Cortex_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Frontal_Cortex_BA9_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Hippocampus_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Nucleus_accumbens_basal_ganglia_Analysis_cis-eQTLs.txt',
                         filename + 'Brain_Putamen_basal_ganglia_Analysis_cis-eQTLs.txt',
                         filename + 'Breast_Mammary_Tissue_Analysis_cis-eQTLs.txt',
                         filename + 'Colon_Sigmoid_Analysis_cis-eQTLs.txt',
                         filename + 'Colon_Transverse_Analysis_cis-eQTLs.txt',
                         filename + 'Esophagus_Gastroesophageal_Junction_Analysis_cis-eQTLs.txt',
                         filename + 'Esophagus_Mucosa_Analysis_cis-eQTLs.txt',
                         filename + 'Esophagus_Muscularis_Analysis_cis-eQTLs.txt',
                         filename + 'Heart_Atrial_Appendage_Analysis_cis-eQTLs.txt',
                         filename + 'Heart_Left_Ventricle_Analysis_cis-eQTLs.txt',
                         filename + 'Liver_Analysis_cis-eQTLs.txt',
                         filename + 'Lung_Analysis_cis-eQTLs.txt',
                         filename + 'Muscle_Skeletal_Analysis_cis-eQTLs.txt',
                         filename + 'Ovary_Analysis_cis-eQTLs.txt',
                         filename + 'Pancreas_Analysis_cis-eQTLs.txt',
                         filename + 'Pituitary_Analysis_cis-eQTLs.txt',
                         filename + 'Prostate_Analysis_cis-eQTLs.txt',
                         filename + 'Skin_Not_Sun_Exposed_Suprapubic_Analysis_cis-eQTLs.txt',
                         filename + 'Skin_Sun_Exposed_Lower_leg_Analysis_cis-eQTLs.txt',
                         filename + 'Spleen_Analysis_cis-eQTLs.txt',
                         filename + 'Stomach_Analysis_cis-eQTLs.txt',
                         filename + 'Testis_Analysis_cis-eQTLs.txt',
                         filename + 'Thyroid_Analysis_cis-eQTLs.txt',
                         filename + 'Uterus_Analysis_cis-eQTLs.txt',
                         filename + 'Vagina_Analysis_cis-eQTLs.txt',
                         filename + 'Whole_Blood_Analysis_cis-eQTLs.txt',
                         filename + 'Small_Intestine_Terminal_Ileum_Analysis_cis-eQTLs.txt'
                         ]
        tissue_list = ['adrenal gland',
                       'blood vessel',
                       'brain',
                       'breast',
                       'colon',
                       'esophagus',
                       'heart',
                       'liver',
                       'lung',
                       'skeletal muscle',
                       'ovary',
                       'pancreas',
                       'pituitary',
                       'prostate',
                       'skin',
                       'spleen',
                       'stomach',
                       'testicle',
                       'thyroid',
                       'uterus',
                       'vagina',
                       'blood',
                       'adrenal gland',
                       'small intestine'
                       ]
        for index in range(0, len(filename_list)):
            with open(filename_list[index], 'r') as f:
                tissue = tissue_list[index]
                next(f)
                for line in f:
                    line = line.split()

                    ensembl = line[0]
                    SNP = line[1]

                    SNP = SNP.split('_')
                    chr = SNP[0]
                    pos = SNP[1]

                    re = 'chr' + chr + '_' + pos  # chr1_618463
                    gene = EnsemblID2geneSymbol(ensembl)
                    w.write(str(gene) + '\t' + str(re) + '\t' + str(tissue))
                    count += 1
                    if count % 1000 == 0:
                        print('line : ', count)
                    w.write('\n')

    f.close()
    w.close()


def import_database_GTEx(filename = './LiCO/data/GTEx_small_intestine_mini.txt'):
    gene_tissue = dict()
    tissue_gene = dict()

    gene_re = dict()
    re_gene = dict()

    re_tissue = dict()
    tissue_re = dict()

    with open(filename, 'r') as f:
        for line in f:
            line = line.split()  # gene    re    tissue
            gene = line[0]
            if gene == 'nan' or gene == 'None':
                continue
            re = line[1]
            tissue = line[2]
            for index in range(3, len(line)):  # organs名字可能是'xxx xxx xxx'形式
                if line[index].isalpha():
                    tissue += ' '
                    tissue += line[index]
                else:
                    break

            if gene not in gene_tissue:
                gene_tissue[gene] = set()
            gene_tissue[gene].add(tissue)

            if gene not in gene_re:
                gene_re[gene] = set()
            gene_re[gene].add(re)

            if tissue not in tissue_gene:
                tissue_gene[tissue] = set()
            tissue_gene[tissue].add(gene)

            if tissue not in tissue_re:
                tissue_re[tissue] = set()
            tissue_re[tissue].add(re)

            if re not in re_gene:
                re_gene[re] = set()
            re_gene[re].add(gene)

            if re not in re_tissue:
                re_tissue[re] = set()
            re_tissue[re].add(tissue)

    return {'gene_tissue': gene_tissue, 'gene_re': gene_re, 'tissue_gene': tissue_gene, 'tissue_re': tissue_re, 're_gene': re_gene, 're_tissue': re_tissue}


def import_database_GTEx_small_intestine(filename = './LiCO/data/GTEx_tissue_spe_eQTL/Small_Intestine_Terminal_Ileum_Analysis_cis-eQTLs.txt'):
    with open('./LiCO/data/GTEx_small_intestine.txt', 'w') as w:
        w.write('gene' + '\t' + 'pos' + '\t' + 'tissue') # 将tissue放最后 方便拿 否则tissue name 和 gene name不好区分开
        w.write('\n')
        count = 0
        with open(filename, 'r') as f:
            tissue = 'small intestine'
            next(f)
            for line in f:
                line = line.split()

                ensembl = line[0]
                SNP = line[1]

                SNP = SNP.split('_')
                chr = SNP[0]
                pos = SNP[1]

                re = 'chr' + chr + '_' + pos  # chr1_618463
                gene = EnsemblID2geneSymbol(ensembl)
                w.write(str(gene) + '\t' + str(re) + '\t' + str(tissue))
                count += 1
                if count % 1000 == 0:
                    print('small_intestine line : ', count)
                w.write('\n')

def import_database_cellTaxonomy(filename = './LiCO/data/Cell_Taxonomy_resource.txt'):
    print('start')
    with open(filename, 'r') as f:
        cnt_line = 0
        for line in f:
            cnt_line += 1
            print('line:', cnt_line)
            line = line.split()
            Species = line[0] + ' ' + line[1]
            print('Species:', Species)
            if Species == 'Homo sapiens':

                # 取tissue
                if line[3] != "NA":
                    tissue = line[3]
                    nextIndex = 4
                    for index in range(4, len(line)):
                        if line[index].isalpha():
                            tissue += ' '
                            tissue += line[index]
                            nextIndex = index + 1
                        else:
                            break
                    print('tissue:', tissue)
                # 取cell type
                cellType = line[nextIndex + 1]
                for index in range(nextIndex + 2, len(line)):
                    if line[index].isalpha():
                        cellType += ' '
                        cellType += line[index]
                    else:
                        break
                print('celltype:', cellType)

                # 取cell marker 只有1个
                cellMarker = line[index + 1]
                print('cellMarker:', cellMarker)

                # cell marker后有Gene_ENTREZID	Gene_Alias	Gene_Ensembl_ID	Uniprot	PFAM	GO2
                # Gene_ENTREZID = line[index + 2]
                # Gene_Alias = line[index + 3]
                # Gene_Ensembl_ID = line[index + 4]
                # Uniprot = line[index + 5]
                # PFAM = line[index + 6]
                # GO2 = line[index + 7]
                # 取生理/病理状态
                condition = line[index + 8]
                print('condition:', condition)

                # 取疾病对应ID 先不转换为疾病名称
                DO = line[index + 9]
                print('DO:', DO)
    f.close()

def import_database_cellTaxonomy_excel(filename = './LiCO/data/Cell_Taxonomy_resource.xlsx'):
    print('start')
    wb2 = load_workbook(filename)  # 读取excel表格
    sheet_ranges = wb2[wb2.sheetnames[0]]  # 定位到表格第一张表

    # dict存储
    gene_dict = dict()  # {'tissue','cellType', 'condition', 'DO'}
    tissue_dict = dict()
    cellType_dict = dict()
    condition_dict = dict()
    DO_dict = dict()

    cnt_line = 0
    for row in sheet_ranges.rows:  # 循环遍历行
        # cnt_line += 1
        # print('line:', cnt_line)

        Species = row[0].value
        # print('Species:', Species)
        if Species == 'Homo sapiens':

            # 取tissue
            tissue = row[2].value
            tissue = tissue.lower()
            if tissue not in tissue_dict:
                tissue_dict[tissue] = {'gene': set(), 'cellType': set(), 'condition': set(), 'DO': set()}
            # print('tissue:', tissue)

            # 取cell type
            cellType = row[4].value
            if cellType not in cellType_dict:
                cellType_dict[cellType] = {'gene': set(), 'tissue': set(), 'condition': set(), 'DO': set()}

            tissue_dict[tissue]['cellType'].add(cellType)
            cellType_dict[cellType]['tissue'].add(tissue)

            # 取cell marker 只有1个
            cellMarker = row[6].value
            if cellMarker not in gene_dict:
                gene_dict[cellMarker] = {'cellType': set(), 'tissue': set(), 'condition': set(), 'DO': set()}
            gene_dict[cellMarker]['tissue'].add(tissue)
            gene_dict[cellMarker]['cellType'].add(cellType)

            tissue_dict[tissue]['gene'].add(cellMarker)
            cellType_dict[cellType]['gene'].add(cellMarker)


            # 取生理/病理状态
            condition = row[13].value
            if condition not in condition_dict:
                condition_dict[condition] = {'cellType': set(), 'tissue': set(), 'gene': set(), 'DO': set()}
            condition_dict[condition]['tissue'].add(tissue)
            condition_dict[condition]['cellType'].add(cellType)
            condition_dict[condition]['gene'].add(cellMarker)

            tissue_dict[tissue]['condition'].add(condition)
            cellType_dict[cellType]['condition'].add(condition)
            gene_dict[cellMarker]['condition'].add(condition)

            # 取疾病对应ID 先不转换为疾病名称
            DO = row[14].value
            if DO not in DO_dict:
                DO_dict[DO] = {'cellType': set(), 'tissue': set(), 'gene': set(), 'condition': set()}
            DO_dict[DO]['tissue'].add(tissue)
            DO_dict[DO]['cellType'].add(cellType)
            DO_dict[DO]['gene'].add(cellMarker)
            DO_dict[DO]['condition'].add(condition)

            tissue_dict[tissue]['DO'].add(DO)
            cellType_dict[cellType]['DO'].add(DO)
            gene_dict[cellMarker]['DO'].add(DO)
            condition_dict[condition]['DO'].add(DO)

    # print('gene Num: ', len(gene_dict.keys()))
    # print('tissue Num: ', len(tissue_dict.keys()))
    # print('cellType Num: ', len(cellType_dict.keys()))
    # print('condition Num: ', len(condition_dict.keys()))
    # print('DO Num: ', len(DO_dict.keys()))

    return {'genes': gene_dict, 'tissues': tissue_dict, 'cellTypes': cellType_dict, 'conditions': condition_dict, 'DOs': DO_dict}





REFERENCE_DICT_DIR = "./LiCO/data/geneID_geneName.txt"

def EnsemblID2geneSymbol(EnsemblID: str = "") -> str:
    EnsemblID = EnsemblID.split(".")[0]  # 去掉尾部类似.1 .2 等的version号
    reference_dict = pd.read_csv(REFERENCE_DICT_DIR, sep='\t')
    df = reference_dict[reference_dict["Gene stable ID"] == EnsemblID]
    try:
        geneSymbol = df["Gene name"].tolist()[0]
        return geneSymbol
    except IndexError:  # 查找不到则返回None
        return None


def geneSymbol2EnsemblID(geneSymbol: str = "") -> str:
    reference_dict = pd.read_csv(REFERENCE_DICT_DIR, sep='\t')
    df = reference_dict[reference_dict["Gene name"] == geneSymbol]
    try:
        EnsemblID = df["Gene stable ID"].tolist()[0]
        return EnsemblID
    except IndexError:  # 查找不到则返回None
        return None



if __name__=='__main__':
    # importppi()
    # importgrnmgi()
    # importgrnhg()
    # importgene_organs()
    # importbodyPartLinks()
    # import_database_cellMarker()
    # import_database_RNA_Atlas()
    # import_database_TRRUST()
    # import_database_grnhg()
    # import_database_hTFtarget()
    # read_GTEx_2_dict()
    # import_database_GTEx_small_intestine()
    # import_predict_b_re_fmax()
    # import_database_cellTaxonomy_excel()
    w = open('./AAAA.txt', 'w')
